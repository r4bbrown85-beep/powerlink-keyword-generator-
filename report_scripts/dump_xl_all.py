import sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl
from pathlib import Path

BASE = Path(r"C:\Users\Administrator\Desktop\0.팀 운영 관련 업무\26년 상반기 및 하반기 업무 정리 보고(65)\암호해제")

files = [
    "미디어채널2팀 팀원들 의견 취합파일_로우데이터 (1).xlsx",
    "미디어채널2팀_업무보고_네이버&메타&토스&당근.xlsx",
    "미디어채널2팀_업무상세_메타&토스_2604 (2).xlsx",
    "네이버 SA 운영 현황 및 2026년 운영 플랜_수정중_0128.xlsx",
]

for fname in files:
    fpath = BASE / fname
    print(f"\n{'='*70}")
    print(f"[파일] {fname}")
    print('='*70)
    try:
        wb = openpyxl.load_workbook(fpath, data_only=True)
        print(f"시트 목록: {wb.sheetnames}")
        for ws in wb.worksheets:
            print(f"\n--- 시트: {ws.title} ({ws.max_row}행 x {ws.max_column}열) ---")
            count = 0
            for row in ws.iter_rows(values_only=True):
                if any(c is not None and str(c).strip() for c in row):
                    cells = [str(c) if c is not None else '' for c in row]
                    cells = [c for c in cells if c.strip()]
                    print(f"  {' | '.join(cells[:10])}")
                    count += 1
                    if count >= 200:
                        print(f"  ... (이하 생략, 총 {ws.max_row}행)")
                        break
    except Exception as e:
        print(f"  오류: {e}")
