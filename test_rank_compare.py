# -*- coding: utf-8 -*-
"""
test_rank_compare.py

키워드별 순위별 성과를 우리 프로그램으로 뽑아서
외부 솔루션(네이버스/에이스퀘어)과 비교할 수 있도록 출력.

실행:
    python test_rank_compare.py

출력:
    콘솔 + output/rank_compare_결과.xlsx
"""
import os
import time
from dotenv import load_dotenv

load_dotenv()

API_KEY     = os.getenv("NAVER_API_KEY", "").strip()
SECRET_KEY  = os.getenv("NAVER_SECRET_KEY", "").strip()
CUSTOMER_ID = os.getenv("NAVER_CUSTOMER_ID", "").strip()

from modules.naver_estimate import get_rank_based_estimates

# ── 테스트 키워드 ──────────────────────────────────────────────────────────────
TEST_KEYWORDS = [
    "강남 임플란트",
    "실손보험 비교",
    "헬로키티 케이크",
    "인테리어 견적",
    "강남 피부과 추천",
    "법인세 신고 대행",
    "남자 헤어스타일",
    "제주도 펜션 추천",
    "파이썬 학원",
    "드리미 로봇청소기",
]

TARGET_RANKS = [1, 2, 3, 4, 5]


def run():
    rows = []

    for kw in TEST_KEYWORDS:
        print(f"\n▶ {kw}")
        t0 = time.time()
        result = get_rank_based_estimates(
            kw, API_KEY, SECRET_KEY, CUSTOMER_ID,
            target_ranks=TARGET_RANKS
        )
        elapsed = time.time() - t0
        print(f"  ({elapsed:.1f}초)")

        for device in ["PC", "MO"]:
            d = result.get(device, {})
            if not d:
                # Fallback
                for rank in TARGET_RANKS:
                    rows.append({
                        "키워드": kw, "디바이스": device, "순위": rank,
                        "예상입찰가": 0, "예상노출": 0,
                        "예상클릭": 0, "예상비용": 0,
                        "CTR": 0, "CPC": 0, "비고": "Fallback"
                    })
                print(f"  [{device}] Fallback (데이터 없음)")
                continue

            for rank in TARGET_RANKS:
                e = d.get(rank)
                if not e:
                    rows.append({
                        "키워드": kw, "디바이스": device, "순위": rank,
                        "예상입찰가": 0, "예상노출": 0,
                        "예상클릭": 0, "예상비용": 0,
                        "CTR": 0, "CPC": 0, "비고": "순위데이터없음"
                    })
                    continue

                rows.append({
                    "키워드":    kw,
                    "디바이스": device,
                    "순위":     rank,
                    "예상입찰가": e["bid"],
                    "예상노출":  e["impressions"],
                    "예상클릭":  e["clicks"],
                    "예상비용":  e["cost"],
                    "CTR":      round(e["ctr"] * 100, 2),
                    "CPC":      e["cpc"],
                    "비고":     "",
                })
                print(f"  [{device}] {rank}위 | 입찰가 {e['bid']:>7,}원 | "
                      f"노출 {e['impressions']:>7,} | 클릭 {e['clicks']:>5,} | "
                      f"비용 {e['cost']:>8,}원 | CTR {e['ctr']*100:.2f}%")

    # ── 엑셀 출력 ──────────────────────────────────────────────────────────────
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from datetime import datetime

        os.makedirs("output", exist_ok=True)
        fname = f"output/rank_compare_{datetime.now().strftime('%m%d_%H%M')}.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "순위별 성과 비교"

        headers = ["키워드", "디바이스", "순위", "예상입찰가", "예상노출",
                   "예상클릭", "예상비용", "CTR(%)", "CPC(원)", "비고",
                   "← 외부솔루션 입찰가", "← 외부솔루션 노출",
                   "← 외부솔루션 클릭", "← 외부솔루션 비용"]

        # 헤더
        header_fill = PatternFill("solid", fgColor="1F4E79")
        ext_fill    = PatternFill("solid", fgColor="375623")
        thin = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font      = Font(bold=True, color="FFFFFF", size=10)
            cell.fill      = ext_fill if "외부솔루션" in h else header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border    = thin

        ws.row_dimensions[1].height = 30

        # 데이터
        pc_fill = PatternFill("solid", fgColor="DEEAF1")
        mo_fill = PatternFill("solid", fgColor="E2EFDA")
        fb_fill = PatternFill("solid", fgColor="FCE4D6")

        for row_idx, r in enumerate(rows, 2):
            vals = [
                r["키워드"], r["디바이스"], r["순위"],
                r["예상입찰가"], r["예상노출"], r["예상클릭"],
                r["예상비용"], r["CTR"], r["CPC"], r["비고"],
                "", "", "", ""  # 외부솔루션 수동 입력칸
            ]
            is_fallback = r["비고"] in ("Fallback", "순위데이터없음")
            row_fill = fb_fill if is_fallback else (pc_fill if r["디바이스"] == "PC" else mo_fill)

            for col, val in enumerate(vals, 1):
                cell = ws.cell(row=row_idx, column=col, value=val)
                cell.border    = thin
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if col <= 10:
                    cell.fill = row_fill
                # 숫자 포맷
                if col in (4, 6, 7, 9):
                    cell.number_format = "#,##0"
                if col == 8:
                    cell.number_format = "0.00"

        # 열 너비
        col_widths = [18, 8, 6, 12, 10, 10, 12, 8, 10, 12, 16, 14, 14, 14]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[ws.cell(1, i).column_letter].width = w

        ws.freeze_panes = "D2"
        wb.save(fname)
        print(f"\n✅ 엑셀 저장: {fname}")
        print("   외부솔루션 컬럼(K~N)에 직접 값 입력 후 비교하세요.")

    except Exception as e:
        print(f"\n엑셀 저장 실패: {e}")
        print("\n── 콘솔 전체 결과 ──")
        print(f"{'키워드':<20} {'디바이스':<6} {'순위':<4} {'입찰가':>8} {'노출':>8} {'클릭':>6} {'비용':>10} {'CTR':>6}")
        print("-" * 80)
        for r in rows:
            print(f"{r['키워드']:<20} {r['디바이스']:<6} {r['순위']:<4} "
                  f"{r['예상입찰가']:>8,} {r['예상노출']:>8,} {r['예상클릭']:>6,} "
                  f"{r['예상비용']:>10,} {r['CTR']:>5.2f}%")


if __name__ == "__main__":
    run()