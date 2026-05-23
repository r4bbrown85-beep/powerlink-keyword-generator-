# -*- coding: utf-8 -*-
"""
main_multi.py
멀티 브랜드 제안서 생성 (브랜드별 시트 구분)
실행: python main_multi.py
     python main_multi.py --profile data/dreame_client_profile.json
"""
import json
import os
import re
import argparse
from datetime import datetime
from dotenv import load_dotenv

from modules.ai_keyword_generator import generate_ai_keyword_plan
from modules.naver_suggest import get_naver_suggestions
from modules.google_suggest import get_google_suggestions
from modules.keyword_filter import filter_ad_keywords, filter_rows_by_brand_context, _STANDALONE_GENERIC_NOUNS
from modules.keyword_scorer import score_keywords
from modules.keyword_type_classifier import classify_keyword_type
from modules.recommendation_engine import build_recommended_keywords, calc_recommendation_score
from modules.summary_builder import build_summary_text
from modules.excel_writer import save_proposal_excel
from modules.naver_keyword_api import get_keyword_stats, get_related_keywords, get_related_keywords_multi
from modules.keyword_normalizer import normalize_keyword_for_ad, normalize_keyword_for_proposal
from modules.bid_simulator import optimize_budget, simulate_expanded, simulate_scenarios, simulate_unconstrained_optimal
from modules.ai_keyword_generator import generate_sa_strategy_memo
from modules.naver_stats import get_registered_keywords, get_keyword_actual_stats
from modules.naver_estimate import get_exposure_min_bid_batch

load_dotenv()

SUGGEST_SEED_LIMIT       = 30
SUGGEST_PER_SOURCE_LIMIT = 5
RELATED_SEED_LIMIT       = 12
RELATED_PER_SEED_LIMIT   = 12

# 카테고리별 키워드 수 상한 (제안서 품질 제어)
CATEGORY_KW_CAPS = {
    "브랜드":  30,
    "일반":    70,
    "경쟁사":  45,
    "상품":    40,
}

AMBIGUOUS_BLACKLIST = [
    "관람평", "영화", "출연진", "감독", "예매", "상영", "시청", "평점",
    "드라마", "ost", "티저", "예고편", "개봉", "흥행",
    "넷플릭스", "왓챠", "티빙", "웨이브",
]

# 업종 불일치 도메인 필터
# 형식: (키워드에_포함된_패턴, [광고주_도메인에_이_단어_중_하나가_있으면_허용])
# 광고주의 category + products + campaign_goal 에서 허용 단어를 확인함
_DOMAIN_MISMATCH_RULES = [
    # ─ 금융·신용 조회 서비스 ─
    ("신용평가",    ["신용", "금융", "신용조회", "평가원", "크레딧"]),
    ("신용등급",    ["신용", "금융", "신용조회", "크레딧"]),
    ("등급확인서",  ["신용", "금융", "자격", "등급"]),
    ("기업신용",    ["신용", "금융", "크레딧"]),
    # ─ ISP·인터넷 회선 서비스 ─
    ("기업인터넷",  ["인터넷", "통신", "isp", "internet", "네트워크"]),
    ("인터넷회선",  ["인터넷", "통신", "isp", "internet"]),
    ("광랜",        ["인터넷", "통신", "광랜", "broadband"]),
    ("인터넷설치",  ["인터넷", "통신", "설치"]),
    # ─ 부동산 ─
    ("부동산",      ["부동산", "임대", "분양", "공인중개", "주택"]),
    ("아파트분양",  ["부동산", "분양", "주택"]),
    # ─ 채용·구직 (HR·인재관리 솔루션 제외) ─
    ("채용공고",    ["채용", "구인", "hr", "인재", "헤드헌팅", "ats"]),
    ("구인구직",    ["채용", "구인", "hr", "인재"]),
    # ─ 여행·관광 ─
    ("호텔예약",    ["호텔", "숙박", "여행", "관광", "예약플랫폼"]),
    ("항공권",      ["항공", "여행", "관광"]),
    # ─ 주식·투자 ─
    ("주식투자",    ["주식", "증권", "투자", "펀드"]),
    ("코인투자",    ["코인", "가상화폐", "블록체인", "투자"]),
    # ─ 교육·훈련 과정 (IT교육 전문 브랜드 제외) ─
    ("양성과정",    ["교육", "훈련", "직업", "인재", "학원", "아카데미"]),
    ("개발자양성",  ["교육", "훈련", "직업", "인재", "학원", "아카데미"]),
    ("교육과정",    ["교육", "훈련", "직업", "인재", "학원", "아카데미", "lms"]),
    ("부트캠프",    ["교육", "훈련", "부트캠프", "코딩"]),
    # ─ 정부지원·창업 프로그램 ─
    ("1인창조기업", ["창업지원", "중소기업", "소상공인", "창업"]),
    ("창조기업",    ["창업지원", "중소기업", "소상공인"]),
    ("사업계획서",  ["창업지원", "스타트업", "투자", "컨설팅", "창업"]),
    ("지원사업",    ["지원사업", "정부지원", "중소기업", "창업지원"]),
    ("사업화지원",  ["지원사업", "정부지원", "중소기업"]),
]


# keyword_filter._STANDALONE_GENERIC_NOUNS를 공유 사용 (중복 정의 방지)
_ULTRA_GENERIC_STANDALONE = _STANDALONE_GENERIC_NOUNS


def _is_ambiguous_keyword(keyword):
    kw = str(keyword).lower()
    return any(term in kw for term in AMBIGUOUS_BLACKLIST)


def _build_advertiser_domain_text(profile):
    """광고주의 업종 컨텍스트 텍스트 — 도메인 불일치 판단용."""
    parts = [
        profile.get("category", ""),
        " ".join(profile.get("products", [])),
        profile.get("campaign_goal", ""),
        " ".join(profile.get("general_keyword_themes", [])),
    ]
    return " ".join(parts).lower().replace(" ", "")


def _is_domain_mismatch(keyword, advertiser_domain_text):
    """
    광고주 업종과 다른 도메인의 키워드인지 확인.
    키워드에 특정 도메인 패턴이 포함되어 있고, 광고주 업종이 해당 도메인과
    무관한 경우 True 반환 → 제거 대상.
    """
    kw = keyword.lower().replace(" ", "")
    for pattern, allowing_domains in _DOMAIN_MISMATCH_RULES:
        if pattern in kw:
            if not any(d in advertiser_domain_text for d in allowing_domains):
                return True
    return False


def _is_too_generic(keyword, all_brand_norms):
    """
    광고 키워드로 쓸 수 없는 초광범위 키워드 여부 확인.
    (1) 2글자 이하 순수 한글, (2) 명시 차단 목록 단독 검색어.
    브랜드명 포함 시 예외.
    """
    kw_no_space = keyword.replace(" ", "")
    kw_norm = normalize_keyword_for_ad(keyword)

    # 브랜드명이 포함된 키워드는 항상 허용
    if any(b and b in kw_norm for b in all_brand_norms):
        return False

    # 명시 차단 목록 — 단독 키워드일 때만 적용
    if kw_no_space in _ULTRA_GENERIC_STANDALONE:
        return True

    # 2글자 이하 순수 한글
    if len(kw_no_space) > 2:
        return False
    if not all('가' <= c <= '힣' for c in kw_no_space):
        return False
    return True


def load_multi_profile(path="data/client_profile.json"):
    with open(path, "r", encoding="utf-8") as f:
        profile = json.load(f)
    return profile


def make_brand_profile(client_profile, brand_cfg):
    """멀티 브랜드 프로파일에서 단일 브랜드 프로파일 생성."""
    return {
        "client":                 client_profile.get("client", ""),
        "brand_name":             brand_cfg.get("brand_name", ""),
        "brand_key":              brand_cfg.get("brand_key", ""),
        "category":               brand_cfg.get("category", ""),
        "brand_urls":             client_profile.get("brand_urls", []),
        "brand_variants":         brand_cfg.get("brand_variants", []),
        "typo_variants":          brand_cfg.get("typo_variants", []),
        "products":               brand_cfg.get("products", []),
        "competitors":            brand_cfg.get("competitors", []),
        "celebrities":            brand_cfg.get("celebrities", []),
        "must_keywords":          brand_cfg.get("must_keywords", []),
        "product_lines":          brand_cfg.get("product_lines", []),
        "general_keyword_themes": brand_cfg.get("general_keyword_themes", []),
        "keyword_categories":     brand_cfg.get("keyword_categories", []),
        "monthly_budget":         client_profile.get("monthly_budget", 5_000_000),
        "sales_channels":         client_profile.get("sales_channels", []),
        "campaign_goal":          client_profile.get("campaign_goal", ""),
        "exclude_keywords":       brand_cfg.get("exclude_keywords",
                                  client_profile.get("exclude_keywords", [])),
        "doc_context":            brand_cfg.get("doc_context", ""),
        "campaign_notes":         brand_cfg.get("campaign_notes", ""),
    }


def uniq_by_ad_keyword(items):
    seen, out = set(), []
    for x in items:
        x = str(x).strip()
        if not x:
            continue
        key = normalize_keyword_for_ad(x)
        if key not in seen:
            out.append(x)
            seen.add(key)
    return out


def chunk_list(items, size):
    return [items[i:i+size] for i in range(0, len(items), size)]


def normalize_seed_keyword(kw):
    kw = str(kw).strip()
    parts = kw.split()
    if len(parts) > 3:
        kw = " ".join(parts[:3])
    weak_tail = ["후기","리뷰","추천","비교","정품","공식몰","지속력","매장위치","시향","구매"]
    parts = kw.split()
    if len(parts) >= 2 and parts[-1] in weak_tail:
        kw = " ".join(parts[:-1]).strip()
    return kw.strip()


def build_rows_from_ai_plan(plan, profile):
    rows = []
    for cat, kws in plan.get("keywords_by_category", {}).items():
        kws = uniq_by_ad_keyword(kws)
        for kw in kws:
            # 네이버 SA는 공백 무시 → 붙여쓰기로 통일
            kw_norm = normalize_keyword_for_proposal(kw)
            if not kw_norm:
                continue
            rows.append({
                "keyword":      kw_norm,
                "category":     cat,
                "keyword_type": classify_keyword_type(kw_norm, profile),
                "score":        50,
                "source":       "ai_seed",
            })
    return rows


def pick_strong_suggest_seeds(rows, profile):
    seeds = []
    seeds.append(profile.get("brand_name", ""))
    for v in profile.get("brand_variants", []):
        seeds.append(v)
    for v in profile.get("typo_variants", []):
        seeds.append(v)
    for p in profile.get("products", []):
        seeds.append(p)
        if profile.get("brand_name"):
            seeds.append(f"{profile.get('brand_name')} {p}")

    # 경쟁사: 전체 사용 + 카테고리 조합으로 경쟁사 키워드 수 확보
    cat_word = profile.get("category", "").split()[0] if profile.get("category") else ""
    for c in profile.get("competitors", []):
        seeds.append(c)
        if cat_word:
            seeds.append(f"{c} {cat_word}")

    for r in rows:
        kw = r["keyword"]
        if len(kw.split()) <= 3:
            seeds.append(kw)
    seeds = uniq_by_ad_keyword([normalize_seed_keyword(x) for x in seeds if x])
    return seeds[:SUGGEST_SEED_LIMIT]


def pick_related_seeds(rows, profile):
    """
    연관 키워드 조회용 씨드 선정.
    카테고리 + 일반 키워드 위주 → 경쟁사 브랜드명은 제외
    (대형 경쟁사 브랜드를 seed로 쓰면 에어컨/냉장고 등 무관 제품 연관키워드가 대량 유입됨)
    """
    seeds = []

    # 1. 카테고리 키워드 우선 (핵심)
    category = profile.get("category", "")
    for cat_kw in category.replace(",", " ").split():
        cat_kw = cat_kw.strip()
        if cat_kw and len(cat_kw) >= 2:
            seeds.append(cat_kw)

    # 1-2. 카테고리 + 주요 수식어 조합
    cat_word = category.split()[0] if category else ""
    if cat_word:
        for modifier in ["추천", "비교", "가격", "순위"]:
            seeds.append(f"{cat_word} {modifier}")

    # 2. AI가 생성한 일반 키워드 중 짧은 것 (2단어 이하)
    general_rows = [r for r in rows
                    if "일반" in r.get("category", "")
                    and len(r["keyword"].split()) <= 2
                    and r.get("source") == "ai_seed"]
    for r in general_rows[:6]:
        seeds.append(r["keyword"])

    # 3. 제품명 (카테고리 관련 제품)
    for p in profile.get("products", [])[:3]:
        seeds.append(p)

    # 4. 브랜드명 (인지도 높은 브랜드는 효과 있음)
    seeds.append(profile.get("brand_name", ""))
    # 브랜드+카테고리 조합
    if cat_word and profile.get("brand_name"):
        seeds.append(f"{profile.get('brand_name')} {cat_word}")

    # 경쟁사는 의도적으로 제외: 대형 브랜드(삼성/LG/애플)를 seed로 쓰면
    # 에어컨/냉장고/TV 등 무관 제품 연관키워드가 대량 유입됨

    seeds = uniq_by_ad_keyword([x for x in seeds if x])
    return seeds[:RELATED_SEED_LIMIT]


def expand_with_suggest(rows, profile):
    seeds = pick_strong_suggest_seeds(rows, profile)
    expanded = []
    for kw in seeds:
        try:
            naver = get_naver_suggestions(kw)
        except Exception:
            naver = []
        try:
            google = get_google_suggestions(kw)
        except Exception:
            google = []
        for s in naver[:SUGGEST_PER_SOURCE_LIMIT]:
            expanded.append((s, "naver_suggest"))
        for s in google[:SUGGEST_PER_SOURCE_LIMIT]:
            expanded.append((s, "google_suggest"))
    return expanded


def expand_with_related_keywords(rows, profile):
    """
    연관 키워드 확장.
    씨드를 5개씩 묶어서 네이버 키워드도구 방식으로 한 번에 조회.
    카테고리 키워드 위주 씨드로 풍부한 연관 키워드 확보.
    """
    api_key = os.getenv("NAVER_API_KEY")
    secret  = os.getenv("NAVER_SECRET_KEY")
    cid     = os.getenv("NAVER_CUSTOMER_ID")
    if not api_key or not secret or not cid:
        return []

    seeds = pick_related_seeds(rows, profile)
    if not seeds:
        return []

    # 5개씩 배치로 묶어서 조회 (네이버 키워드도구와 동일한 방식)
    all_related = get_related_keywords_multi(
        seeds, api_key, secret, cid,
        query_type="RELATED", batch_size=5
    )

    # 중복 제거 및 limit 적용
    seen_norm, expanded = set(), []
    for item in all_related:
        norm = normalize_keyword_for_ad(item["keyword"])
        if norm not in seen_norm:
            seen_norm.add(norm)
            expanded.append(item)

    print(f"  연관키워드 확장: {len(expanded)}개 (씨드 {len(seeds)}개)")
    return expanded


def guess_category_for_suggestion(keyword, selected_categories, profile):
    k  = str(keyword).strip()
    nk = normalize_keyword_for_ad(k)
    brand_all = (
        [profile.get("brand_name", "")]
        + profile.get("brand_variants", [])
        + profile.get("typo_variants", [])
    )
    for b in brand_all:
        if b and normalize_keyword_for_ad(b) in nk:
            for c in selected_categories:
                if "브랜드" in c:
                    return c
    for comp in profile.get("competitors", []):
        if comp and normalize_keyword_for_ad(comp) in nk:
            for c in selected_categories:
                if "경쟁사" in c:
                    return c
    for p in profile.get("products", []):
        if p and normalize_keyword_for_ad(p) in nk:
            for c in selected_categories:
                if "상품" in c or "제품" in c:
                    return c
    for c in selected_categories:
        if "일반" in c or "카테고리" in c:
            return c
    return selected_categories[0] if selected_categories else "일반 키워드"


def build_relevance_filter(profile):
    tokens = set()
    def add_tokens(text):
        if not text:
            return
        t = str(text).strip().lower()
        tokens.add(t)
        tokens.add(t.replace(" ", ""))
        # 2글자 이상 단어 단위로도 추가
        for word in t.split():
            if len(word) >= 2:
                tokens.add(word)

    add_tokens(profile.get("brand_name", ""))
    add_tokens(profile.get("category", ""))
    for v in profile.get("brand_variants", []):
        add_tokens(v)
    for v in profile.get("typo_variants", []):
        add_tokens(v)
    for p in profile.get("products", []):
        add_tokens(p)
    for c in profile.get("competitors", []):
        add_tokens(c)
    for kw in profile.get("must_keywords", []):
        if isinstance(kw, dict):
            add_tokens(kw.get("keyword", ""))
        else:
            add_tokens(kw)

    # 카테고리별 일반 토큰
    cat = profile.get("category", "").lower()
    cat_tokens_map = {
        "헤어드라이기": ["헤어드라이기","헤어드라이어","드라이기","드라이어","헤어케어","고속드라이","헤어"],
        "물걸레청소기": ["물걸레","청소기","물걸레청소기","스팀청소기","바닥청소","무선물걸레"],
        "로봇청소기":   ["로봇청소기","로봇","청소기","자동청소","물걸레로봇","로보락"],
        "무선청소기":   ["무선청소기","스틱청소기","핸디청소기","청소기","무선","스틱"],
        "향수":         ["향수","퍼퓸","니치","명품","오드퍼퓸","프래그런스"],
        "노트북":       ["노트북","laptop","울트라북","맥북","삼성노트북","lg노트북","게이밍노트북","학생노트북","사무용노트북","노트북추천"],
        "태블릿":       ["태블릿","tablet","아이패드","갤럭시탭","안드로이드"],
        "스마트폰":     ["스마트폰","휴대폰","아이폰","갤럭시","핸드폰","폰"],
        "에어컨":       ["에어컨","냉방","인버터","벽걸이","스탠드에어컨"],
        "냉장고":       ["냉장고","김치냉장고","양문형","냉동"],
        "세탁기":       ["세탁기","드럼세탁기","통돌이","건조기"],
        "TV":           ["tv","텔레비전","oled","qled","모니터","디스플레이"],
        "카메라":       ["카메라","미러리스","dslr","렌즈","촬영"],
        "의류":         ["옷","의류","패션","코트","자켓","티셔츠","원피스"],
        "화장품":       ["화장품","스킨케어","로션","세럼","크림","선크림","마스크팩"],
        "건강식품":     ["건강식품","영양제","프로바이오틱스","비타민","보충제"],
        "식품":         ["식품","음식","반찬","간식","음료","커피","차"],
        "가구":         ["가구","소파","침대","책상","의자","장롱"],
        "외장하드":       ["외장하드","외장ssd","외장 ssd","hdd","ssd","pssd","포터블ssd","저장장치","usb드라이브"],
        "외장ssd":        ["외장ssd","외장 ssd","외장하드","hdd","ssd","pssd","포터블","저장장치"],
        "저장장치":       ["저장장치","외장하드","ssd","hdd","외장ssd","pssd","포터블"],
    }
    for cat_key, toks in cat_tokens_map.items():
        if cat_key in cat:
            for t in toks:
                tokens.add(t)
    return tokens


_SHORT_ASCII_RE_CACHE: dict = {}

def _word_boundary_match(token: str, text: str) -> bool:
    """단어 경계 매칭 — 짧은 영문 토큰이 다른 단어 내부에 포함된 경우 제외.
    예: 'gram' → 'grammar' 불일치, 'lg gram' 내의 'gram' 일치."""
    if token not in _SHORT_ASCII_RE_CACHE:
        _SHORT_ASCII_RE_CACHE[token] = re.compile(
            rf'(?<![a-z]){re.escape(token)}(?![a-z])'
        )
    return bool(_SHORT_ASCII_RE_CACHE[token].search(text))


def is_relevant_keyword(keyword, relevance_tokens, profile):
    kw_lower    = str(keyword).strip().lower()
    kw_no_space = kw_lower.replace(" ", "")
    for token in relevance_tokens:
        if not token:
            continue
        # 짧은 순수 영문 토큰(≤5자)은 단어 경계 매칭 — 'gram' ≠ 'grammar'
        if len(token) <= 5 and token.isascii() and token.isalpha():
            if _word_boundary_match(token, kw_lower) or _word_boundary_match(token, kw_no_space):
                return True
        else:
            if token in kw_lower or token in kw_no_space:
                return True
    return False


def apply_category_caps(rows):
    """카테고리별 키워드 수 상한 적용.
    선발 우선순위: ai_seed 여부 → recommendation_score(효율 점수, 있을 때) → score(기본 점수).
    Naver stats 조회 + recommendation_score 계산 후에 호출해야 효율 기준 선발이 가능함.
    """
    by_cat = {}
    for row in rows:
        cat = row.get("category", "일반 키워드")
        by_cat.setdefault(cat, []).append(row)

    result = []
    for cat, cat_rows in by_cat.items():
        cap = None
        for key, limit in CATEGORY_KW_CAPS.items():
            if key in cat:
                cap = limit
                break

        if cap is None or len(cat_rows) <= cap:
            result.extend(cat_rows)
            continue

        sorted_rows = sorted(cat_rows, key=lambda x: (
            0 if x.get("source") == "ai_seed" else 1,
            # recommendation_score가 있으면(stats 조회 후) 효율 기준으로 선발,
            # 없으면(stats 조회 전 호출 시) 기본 score로 fallback
            -x.get("recommendation_score", x.get("score", 0)),
        ))
        result.extend(sorted_rows[:cap])
        print(f"  카테고리 상한: {cat} {len(cat_rows)}→{cap}개")

    return result


def _is_category_relevant(keyword, profile):
    """카테고리 핵심 단어가 키워드에 포함되어 있는지 확인."""
    kw_lower = keyword.lower().replace(" ", "")
    category = profile.get("category", "").lower()
    for cat_word in category.replace(",", " ").split():
        if len(cat_word) >= 2 and cat_word in kw_lower:
            return True
    for p in profile.get("products", []):
        if p and p.lower().replace(" ", "") in kw_lower:
            return True
    return False


def filter_unrelated_keywords(rows, profile):
    relevance_tokens = build_relevance_filter(profile)
    brand_name_norm  = normalize_keyword_for_ad(profile.get("brand_name", ""))
    brand_variants   = [normalize_keyword_for_ad(v)
                        for v in profile.get("brand_variants", [])
                        + profile.get("typo_variants", [])]
    all_brand_norms  = [b for b in [brand_name_norm] + brand_variants if b]
    exclude_kws      = [normalize_keyword_for_ad(e)
                        for e in profile.get("exclude_keywords", [])]

    # 경쟁사 브랜드명 정규화 목록
    competitor_norms = [normalize_keyword_for_ad(c)
                        for c in profile.get("competitors", []) if c]

    # 경쟁사 키워드 카테고리명 찾기
    comp_cat_name = "경쟁사 키워드"
    for cat in profile.get("keyword_categories", []):
        if cat.get("type") == "competitor":
            comp_cat_name = cat.get("name", "경쟁사 키워드")
            break

    advertiser_domain_text = _build_advertiser_domain_text(profile)

    filtered, removed = [], 0
    removed_reasons: dict[str, int] = {}

    def _drop(reason):
        nonlocal removed
        removed += 1
        removed_reasons[reason] = removed_reasons.get(reason, 0) + 1

    for row in rows:
        source  = row.get("source", "")
        kw      = row.get("keyword", "")
        kw_norm = normalize_keyword_for_ad(kw)

        if kw_norm in exclude_kws:
            _drop("exclude_kws")
            continue

        # ── 전체 키워드 공통 필터 (ai_seed 포함) ─────────────────────────
        # 1. 업종 불일치 도메인 패턴 (신용평가, 기업인터넷 등)
        if _is_domain_mismatch(kw, advertiser_domain_text):
            _drop("domain_mismatch")
            continue

        # 2. 2글자 이하 초광범위 한글 키워드 (기업, 업무 등 단독)
        if _is_too_generic(kw, all_brand_norms):
            _drop("too_generic")
            continue
        # ─────────────────────────────────────────────────────────────────

        # 지역명 + 카테고리 + 수리 패턴 (제주시컴퓨터수리 등) — 확장 키워드만 적용
        if source != "ai_seed" and "수리" in kw.lower():
            has_brand_in_kw = any(b and b in kw_norm for b in all_brand_norms)
            if not has_brand_in_kw:
                _drop("repair_pattern")
                continue

        # profile의 exclude_keywords로 제외 (광고주별 설정)
        exclude_kw_list = profile.get("exclude_keywords", [])
        if exclude_kw_list:
            kw_lower = kw.lower().replace(" ","")
            if any(excl.lower().replace(" ","") in kw_lower for excl in exclude_kw_list):
                _drop("exclude_kw_list")
                continue

        # 경쟁사 키워드 자동 재분류:
        # 경쟁사 브랜드명이 포함된 키워드는 강제로 경쟁사 카테고리로 분류
        # (일반 키워드로 잘못 분류되어 예산 상한 없이 배정되는 문제 방지)
        if source != "ai_seed":
            is_competitor_kw = any(
                comp and comp in kw_norm
                for comp in competitor_norms
            )
            if is_competitor_kw:
                # 경쟁사 브랜드명이 포함되어도 카테고리와 무관하면 제거
                # (삼성에어컨, 삼성냉장고 등 타 제품군 유입 방지)
                if not _is_category_relevant(kw, profile):
                    _drop("competitor_irrelevant")
                    continue
                row = dict(row)
                row["category"] = comp_cat_name
                row["keyword_type"] = "COMPETITOR"

        if source == "ai_seed":
            filtered.append(row)
            continue
        must_kws = []
        for m in profile.get("must_keywords", []):
            if isinstance(m, dict):
                must_kws.append(normalize_keyword_for_ad(m.get("keyword", "")))
            else:
                must_kws.append(normalize_keyword_for_ad(m))
        if kw_norm in must_kws:
            filtered.append(row)
            continue
        if _is_ambiguous_keyword(kw):
            has_brand = any(b in kw_norm for b in all_brand_norms)
            if not has_brand:
                _drop("ambiguous")
                continue
        if is_relevant_keyword(kw, relevance_tokens, profile):
            filtered.append(row)
        else:
            _drop("not_relevant")

    reason_summary = ", ".join(f"{r}:{n}" for r, n in sorted(removed_reasons.items()))
    print(f"  무관련 키워드 제거: {removed}개 [{reason_summary}] / 남은 키워드: {len(filtered)}개")
    return filtered


def merge_all_rows(ai_rows, suggest_items, related_items, selected_categories, profile):
    merged, seen_norm = [], set()
    for row in ai_rows:
        norm = normalize_keyword_for_ad(row["keyword"])
        if norm not in seen_norm:
            merged.append(row)
            seen_norm.add(norm)
    for kw, source in suggest_items:
        kw = normalize_keyword_for_proposal(kw)
        norm = normalize_keyword_for_ad(kw)
        if norm in seen_norm:
            continue
        merged.append({
            "keyword":      kw,
            "category":     guess_category_for_suggestion(kw, selected_categories, profile),
            "keyword_type": classify_keyword_type(kw, profile),
            "score":        50,
            "source":       source,
        })
        seen_norm.add(norm)
    for item in related_items:
        kw = normalize_keyword_for_proposal(item["keyword"])
        norm = normalize_keyword_for_ad(kw)
        if norm in seen_norm:
            continue
        merged.append({
            "keyword":      kw,
            "category":     guess_category_for_suggestion(kw, selected_categories, profile),
            "keyword_type": classify_keyword_type(kw, profile),
            "score":        50,
            "source":       item.get("source", "naver_related"),
            "pc_impr":      item.get("pc_impr", 0),
            "mo_impr":      item.get("mo_impr", 0),
            "pc_click":     item.get("pc_click", 0),
            "mo_click":     item.get("mo_click", 0),
            "competition":  item.get("competition", "MID"),
        })
        seen_norm.add(norm)
    return merged


def attach_naver_stats(rows):
    api_key = os.getenv("NAVER_API_KEY")
    secret  = os.getenv("NAVER_SECRET_KEY")
    cid     = os.getenv("NAVER_CUSTOMER_ID")
    if not api_key or not secret or not cid:
        for row in rows:
            row.setdefault("pc_impr", 0)
            row.setdefault("mo_impr", 0)
            row.setdefault("pc_click", 0)
            row.setdefault("mo_click", 0)
            row.setdefault("competition", "MID")
            row.setdefault("monthlySearchVolume", 0)
            row.setdefault("topOfPageBid", 1000)
        return rows

    need_api = [row for row in rows if not row.get("pc_impr") and not row.get("mo_impr")]
    print(f"  네이버 stats 조회 필요: {len(need_api)}개")

    keywords  = [row["keyword"] for row in need_api]
    stats_map = {}
    for batch in chunk_list(keywords, 5):
        try:
            res = get_keyword_stats(batch, api_key, secret, cid)
            stats_map.update(res)
        except Exception as e:
            print(f"  네이버 API 오류: {e}")

    for row in rows:
        if row.get("pc_impr") or row.get("mo_impr"):
            row.setdefault("monthlySearchVolume",
                           row.get("pc_impr", 0) + row.get("mo_impr", 0))
            row.setdefault("competition", "MID")
        else:
            data               = stats_map.get(row["keyword"], {})
            row["pc_impr"]     = data.get("pc_impr", 0)    # 월간 PC 검색량 (쿼리수)
            row["mo_impr"]     = data.get("mo_impr", 0)    # 월간 MO 검색량 (쿼리수)
            row["pc_click"]    = data.get("pc_click", 0)   # 월간 시장 평균 PC 클릭수
            row["mo_click"]    = data.get("mo_click", 0)   # 월간 시장 평균 MO 클릭수
            row["ctr"]         = data.get("ctr", 0)
            row["competition"] = data.get("competition", "MID")
            row["monthlySearchVolume"] = row["pc_impr"] + row["mo_impr"]
            # plAvgDepth: 파워링크 평균 게재 순위 (있을 때만 저장, bid_simulator fallback에서 활용)
            pl_depth = data.get("pl_avg_depth")
            if pl_depth is not None:
                row["pl_avg_depth"] = pl_depth

        comp   = str(row.get("competition", "MID")).upper()
        volume = row.get("monthlySearchVolume", 0)
        est_top_bid = 1800 if comp == "HIGH" else (600 if comp == "LOW" else 1000)
        if volume >= 30000:   est_top_bid += 500
        elif volume >= 10000: est_top_bid += 300
        elif volume >= 3000:  est_top_bid += 150
        row["topOfPageBid"] = est_top_bid
    return rows


def attach_budget_plan(recommended_rows, total_budget, brand_profile=None):
    keyword_rows = []
    for row in recommended_rows:
        keyword_rows.append({
            "keyword":              row["keyword"],
            "category":             row.get("category", ""),
            "keyword_type":         row.get("keyword_type", "GENERIC"),
            "competition":          row.get("competition", "MID"),
            "monthlySearchVolume":  row.get("monthlySearchVolume", 1000),
            "topOfPageBid":         row.get("topOfPageBid", 1000),
            "recommendation_score": row.get("recommendation_score", 50),
            "pc_impr":              row.get("pc_impr", 0),
            "mo_impr":              row.get("mo_impr", 0),
            "pc_click":             row.get("pc_click", 0),
            "mo_click":             row.get("mo_click", 0),
        })

    selected, total_cost, selected_keywords, standby_rows, all_options_map = \
        optimize_budget(keyword_rows, total_budget,
                        competitors=(brand_profile or {}).get("competitors", []),
                        cat_config=(brand_profile or {}).get("keyword_categories", None))

    selected_map = {sim["keyword"]: sim for sim in selected}

    cat_bid_avg = {}
    for sim in selected:
        if sim.get("is_fallback", False):
            continue
        cat = sim.get("category", "일반 키워드")
        pc_bid = sim.get("pc_bid", sim.get("bid", 0)) or 0
        mo_bid = sim.get("mo_bid", sim.get("bid", 0)) or 0
        if cat not in cat_bid_avg:
            cat_bid_avg[cat] = {"pc": [], "mo": []}
        if pc_bid > 0: cat_bid_avg[cat]["pc"].append(pc_bid)
        if mo_bid > 0: cat_bid_avg[cat]["mo"].append(mo_bid)

    cat_avg_pc = {cat: int(sum(v["pc"])/len(v["pc"])) for cat, v in cat_bid_avg.items() if v["pc"]}
    cat_avg_mo = {cat: int(sum(v["mo"])/len(v["mo"])) for cat, v in cat_bid_avg.items() if v["mo"]}

    # 폴백 키워드: 노출최소입찰가 API로 실제 최소 입찰가 조회
    _fallback_kws = [
        row["keyword"] for row in recommended_rows
        if selected_map.get(row["keyword"], {}).get("is_fallback", False)
    ]
    _api_key = os.getenv("NAVER_API_KEY", "")
    _secret  = os.getenv("NAVER_SECRET_KEY", "")
    _cid     = os.getenv("NAVER_CUSTOMER_ID", "")
    exp_min_pc_map: dict = {}
    exp_min_mo_map: dict = {}
    if _fallback_kws and _api_key and _secret and _cid:
        try:
            exp_min_pc_map = get_exposure_min_bid_batch(_fallback_kws, "PC", _api_key, _secret, _cid)
            exp_min_mo_map = get_exposure_min_bid_batch(_fallback_kws, "MOBILE", _api_key, _secret, _cid)
            print(f"  [폴백] 노출최소입찰가 조회: PC {len(exp_min_pc_map)}개 / MO {len(exp_min_mo_map)}개")
        except Exception as _e:
            print(f"  [폴백] 노출최소입찰가 조회 실패 (무시): {_e}")

    for row in recommended_rows:
        sim = selected_map.get(row["keyword"])
        if sim:
            is_fb = sim.get("is_fallback", False)
            row["proposed_bid"]    = sim.get("bid", "")
            row["proposed_bid_pc"] = sim.get("pc_bid", sim.get("bid", ""))
            row["proposed_bid_mo"] = sim.get("mo_bid", sim.get("bid", ""))
            row["is_fallback"]     = is_fb
            row["not_selected"]    = False

            if is_fb:
                kw = row["keyword"]
                top_bid    = row.get("topOfPageBid", 0) or 0
                exp_min_pc = exp_min_pc_map.get(kw, 0) or 0
                exp_min_mo = exp_min_mo_map.get(kw, 0) or 0
                if exp_min_pc > 0:
                    fallback_pc_bid = max(70, exp_min_pc)
                    fallback_mo_bid = max(70, exp_min_mo if exp_min_mo > 0 else int(exp_min_pc * 0.7))
                elif top_bid > 0:
                    fallback_pc_bid = max(70, int(top_bid * 0.3))
                    fallback_mo_bid = max(70, int(top_bid * 0.2))
                else:
                    fallback_pc_bid = 300
                    fallback_mo_bid = 200
                row["proposed_bid"]    = fallback_pc_bid
                row["proposed_bid_pc"] = fallback_pc_bid
                row["proposed_bid_mo"] = fallback_mo_bid
                row["proposed_rank"] = row["proposed_rank_pc"] = row["proposed_rank_mo"] = ""
                for f in ["sim_impressions","sim_ctr","sim_clicks","sim_cpc","sim_cost","anchor_bid",
                          "pc_sim_impressions","pc_sim_ctr","pc_sim_clicks","pc_sim_cpc","pc_sim_cost",
                          "mo_sim_impressions","mo_sim_ctr","mo_sim_clicks","mo_sim_cpc","mo_sim_cost"]:
                    row[f] = ""
            else:
                row["proposed_rank"]      = sim["rank"]
                row["proposed_rank_pc"]   = sim.get("rank_pc", "")
                row["proposed_rank_mo"]   = sim.get("rank_mo", "")
                row["sim_impressions"]    = sim["impressions"]
                row["sim_ctr"]            = sim["ctr"]
                row["sim_clicks"]         = sim["clicks"]
                row["sim_cpc"]            = sim["cpc"]
                row["sim_cost"]           = sim["cost"]
                row["anchor_bid"]         = sim.get("anchor_bid", "")
                row["pc_sim_impressions"] = sim.get("pc_impressions", 0)
                row["pc_sim_ctr"]         = sim.get("pc_ctr", 0)
                row["pc_sim_clicks"]      = sim.get("pc_clicks", 0)
                row["pc_sim_cpc"]         = sim.get("pc_cpc", 0)
                row["pc_sim_cost"]        = sim.get("pc_cost", 0)
                row["mo_sim_impressions"] = sim.get("mo_impressions", 0)
                row["mo_sim_ctr"]         = sim.get("mo_ctr", 0)
                row["mo_sim_clicks"]      = sim.get("mo_clicks", 0)
                row["mo_sim_cpc"]         = sim.get("mo_cpc", 0)
                row["mo_sim_cost"]        = sim.get("mo_cost", 0)
        else:
            top_bid     = row.get("topOfPageBid", 0) or 0
            competition = str(row.get("competition", "MID")).upper()
            suggest_bid = max(70, int(top_bid * 0.5)) if top_bid > 0 else \
                          {"HIGH": 300, "LOW": 70}.get(competition, 150)
            row["proposed_bid"] = row["proposed_bid_pc"] = row["proposed_bid_mo"] = suggest_bid
            row["proposed_rank"] = row["proposed_rank_pc"] = row["proposed_rank_mo"] = ""
            for f in ["sim_impressions","sim_ctr","sim_clicks","sim_cpc","sim_cost","anchor_bid",
                      "pc_sim_impressions","pc_sim_ctr","pc_sim_clicks","pc_sim_cpc","pc_sim_cost",
                      "mo_sim_impressions","mo_sim_ctr","mo_sim_clicks","mo_sim_cpc","mo_sim_cost"]:
                row[f] = ""
            row["is_fallback"]  = False
            row["not_selected"] = True

    return recommended_rows, total_cost, standby_rows, all_options_map


def run_single_brand(brand_profile, brand_name, progress_cb=None):
    """
    단일 브랜드 키워드 제안 데이터 생성.
    progress_cb: callable(step: str, pct: float) — UI 진행률 콜백 (선택)
    """
    def _progress(step: str, pct: float):
        if progress_cb:
            progress_cb(step, pct)

    total_budget = brand_profile.get("monthly_budget", 5_000_000)
    print(f"\n{'='*50}")
    print(f"브랜드: {brand_name}")
    print(f"{'='*50}")

    # 2-0. SA 전략 브리핑 자동 생성 (캠페인 메모와 별개로 항상 실행)
    _progress("SA 전략 분석 중...", 0.06)
    print("  [2-0] SA 전략 브리핑 자동 생성...")
    sa_memo = generate_sa_strategy_memo(brand_profile)
    if sa_memo:
        brand_profile["sa_strategy_memo"] = sa_memo

    # 2. AI 키워드 생성
    _progress("AI 키워드 생성 중...", 0.14)
    print("  [2] AI 키워드 생성...")
    plan                = generate_ai_keyword_plan(brand_profile)
    selected_categories = plan.get("selected_categories", [])
    category_desc       = plan.get("category_descriptions", {})
    ai_rows             = build_rows_from_ai_plan(plan, brand_profile)
    print(f"  AI 키워드: {len(ai_rows)}개")

    # 3. 자동완성 확장
    _progress("자동완성 키워드 확장 중...", 0.25)
    print("  [3] 자동완성 확장...")
    suggest_items = expand_with_suggest(ai_rows, brand_profile)
    suggest_items = filter_rows_by_brand_context(suggest_items, brand_profile)

    # 3-2. 연관키워드 확장
    _progress("네이버 연관 키워드 수집 중...", 0.38)
    print("  [3-2] 연관키워드 확장...")
    related_items = expand_with_related_keywords(ai_rows, brand_profile)
    related_items = filter_rows_by_brand_context(related_items, brand_profile)

    # 4. 병합 및 필터링
    _progress("키워드 필터링 및 분류 중...", 0.50)
    print("  [4] 병합 및 필터링...")
    rows = merge_all_rows(ai_rows, suggest_items, related_items,
                          selected_categories, brand_profile)
    rows = filter_unrelated_keywords(rows, brand_profile)
    filtered_keywords = filter_ad_keywords([row["keyword"] for row in rows])
    filtered_set      = {normalize_keyword_for_ad(x) for x in filtered_keywords}
    rows              = [row for row in rows
                         if normalize_keyword_for_ad(row["keyword"]) in filtered_set]

    # 4-1. 명시적 제외 키워드 최종 필터
    # AI 캐시·자동완성 재생성으로 제외 키워드가 돌아오는 경우를 차단
    _exc_set = {e.lower().replace(" ", "") for e in brand_profile.get("exclude_keywords", []) if e}
    if _exc_set:
        before = len(rows)
        rows = [r for r in rows if r.get("keyword", "").lower().replace(" ", "") not in _exc_set]
        if before - len(rows):
            print(f"  제외 키워드 최종 필터: {before}개 → {len(rows)}개 ({before - len(rows)}개 제거)")

    # 점수 부여
    keywords  = [row["keyword"] for row in rows]
    scored    = score_keywords(keywords)
    score_map = {k: s for k, s in scored}
    for row in rows:
        row["score"] = score_map.get(row["keyword"], 50)
    print(f"  전체 키워드 (필터 후): {len(rows)}개")

    # 5. 네이버 데이터 조회 (카테고리 상한 적용 전 — 효율 기준 선발을 위해 stats 먼저 조회)
    _progress("네이버 검색 통계 조회 중...", 0.60)
    print("  [5] 네이버 데이터 조회...")
    rows = attach_naver_stats(rows)

    # 5-1. 계정 실집행 데이터 조회 (선택적 — API 권한 없으면 빈 결과)
    _api_key     = os.getenv("NAVER_API_KEY", "")
    _secret      = os.getenv("NAVER_SECRET_KEY", "")
    _customer_id = os.getenv("NAVER_CUSTOMER_ID", "")
    registered_kws = set()
    actual_stats   = {}
    if _api_key and _secret and _customer_id:
        try:
            registered_kws = get_registered_keywords(_api_key, _secret, _customer_id)
            actual_stats   = get_keyword_actual_stats(_api_key, _secret, _customer_id)
        except Exception as e:
            print(f"  [실적] 계정 데이터 조회 실패 (무시): {e}")

    # 기등록 여부 + 실측 순위 태깅
    for row in rows:
        kw = row.get("keyword", "")
        row["already_registered"] = kw in registered_kws
        if kw in actual_stats:
            row["actual_avg_rank"] = actual_stats[kw].get("avg_rank", 0)
            row["actual_ctr"]      = actual_stats[kw].get("ctr", 0)
            row["actual_clicks"]   = actual_stats[kw].get("clicks", 0)

    if registered_kws:
        already_cnt = sum(1 for r in rows if r.get("already_registered"))
        print(f"  [계정] 기등록 키워드: {already_cnt}개 / 전체 {len(rows)}개")

    # 6. 추천 키워드 선정
    _progress("추천 키워드 선정 중...", 0.72)
    for row in rows:
        row["recommendation_score"] = calc_recommendation_score(row)

    # 6-1. 카테고리 상한 적용 — recommendation_score 계산 후 실행해 효율 기준으로 선발
    rows = apply_category_caps(rows)
    print(f"  카테고리 상한 후: {len(rows)}개")

    recommended = sorted(rows, key=lambda x: -x.get("recommendation_score", 0))
    print(f"  추천 키워드: {len(recommended)}개")

    # 7. 예산 시뮬레이션
    _progress("예산 시뮬레이션 중...", 0.83)
    print("  [7] 예산 시뮬레이션...")
    recommended, total_cost, standby_rows, all_options_map = \
        attach_budget_plan(recommended, total_budget, brand_profile=brand_profile)
    print(f"  예상 총 비용: {total_cost:,}원")

    # 8. Summary
    summary = build_summary_text(brand_profile, rows, recommended)

    # 8-2. 확장 시뮬레이션
    _progress("확장 시나리오 계산 중...", 0.93)
    print("  [8-2] 확장 시뮬레이션...")
    current_rows  = [r for r in recommended if not r.get("not_selected", False)]
    not_sel_rows  = [r for r in recommended if r.get("not_selected", False)]
    not_sel_kw_rows = [{
        "keyword":             row["keyword"],
        "category":            row.get("category", ""),
        "keyword_type":        row.get("keyword_type", "GENERIC"),
        "competition":         row.get("competition", "MID"),
        "monthlySearchVolume": row.get("monthlySearchVolume", 0),
        "topOfPageBid":        row.get("topOfPageBid", 1000),
        "pc_impr":             row.get("pc_impr", 0),
        "mo_impr":             row.get("mo_impr", 0),
    } for row in not_sel_rows]

    scenario_data = simulate_scenarios(current_rows, not_sel_kw_rows,
                                       total_budget, all_options_map)

    # 예산 무제한 최적 효율 제안 시뮬레이션
    try:
        optimal_data = simulate_unconstrained_optimal(current_rows, not_sel_rows)
    except Exception as _e:
        print(f"  [최적효율] 시뮬레이션 실패 (무시): {_e}")
        optimal_data = None

    return {
        "brand_name":      brand_name,
        "brand_category":  brand_profile.get("category", ""),
        "monthly_budget":  brand_profile.get("monthly_budget", 500000),
        "rows":          rows,
        "recommended":   recommended,
        "category_desc": category_desc,
        "summary":       summary,
        "standby_rows":  standby_rows,
        "scenario_data": scenario_data,
        "optimal_data":  optimal_data,
        "total_cost":    total_cost,
    }


def run_budget_simulation(keywords: list, budget: int, client_name: str,
                          progress_cb=None) -> dict:
    """
    확정 키워드 리스트 + 예산 → 최적 순위/입찰가/예상성과 시뮬레이션.
    AI 키워드 생성 없이 네이버 API 데이터 조회 + 예산 최적화만 실행.
    """
    def _progress(step: str, pct: float):
        if progress_cb:
            progress_cb(step, pct)

    # 기본 row 구조 생성 — 클라이언트명 포함 키워드는 브랜드 키워드로 분류
    client_lower = client_name.lower().replace(" ", "")
    rows = []
    seen = set()
    for kw in keywords:
        kw = str(kw).strip()
        if not kw:
            continue
        norm = normalize_keyword_for_ad(kw)
        if norm in seen:
            continue
        seen.add(norm)
        kw_norm_lower = kw.lower().replace(" ", "")
        if client_lower and len(client_lower) >= 2 and client_lower in kw_norm_lower:
            cat, kw_type = "브랜드 키워드", "BRAND"
        else:
            cat, kw_type = "일반 키워드", "GENERIC"
        rows.append({
            "keyword":      normalize_keyword_for_proposal(kw),
            "category":     cat,
            "keyword_type": kw_type,
            "score":        50,
            "source":       "user_input",
        })

    if not rows:
        return None

    _progress("네이버 검색 통계 조회 중...", 0.15)
    print(f"  [시뮬레이터] 키워드 {len(rows)}개 통계 조회...")
    rows = attach_naver_stats(rows)

    # 점수 및 추천 순위 계산
    _progress("키워드 분석 중...", 0.40)
    keywords_list = [row["keyword"] for row in rows]
    scored     = score_keywords(keywords_list)
    score_map  = {k: s for k, s in scored}
    for row in rows:
        row["score"] = score_map.get(row["keyword"], 50)
        row["recommendation_score"] = calc_recommendation_score(row)

    recommended = sorted(rows, key=lambda x: -x.get("recommendation_score", 0))

    # 예산 시뮬레이션 — 시뮬레이터 전용 카테고리 설정 (max_single_ratio 완화)
    # AI 제안서(15% 캡)와 달리 시뮬레이터는 키워드 수가 적으므로 캡을 넓혀야 예산을 제대로 활용
    # min_rank: 허용 최고 순위 번호. 브랜드는 1위 OK, 일반은 2위 이상으로 제한
    # → upgrade 단계에서 일반 키워드가 1위까지 올라가는 것을 방지
    _sim_cat_config = [
        {"name": "브랜드 키워드", "type": "brand",   "priority": 1.30,
         "min_keywords": 0, "target_rank": 1, "max_rank": 3, "min_rank": 1,
         "max_single_ratio": 1.00, "cpc_factor": 1.10, "color": "BDD7EE"},
        {"name": "일반 키워드",   "type": "general", "priority": 1.00,
         "min_keywords": 0, "target_rank": 3, "max_rank": 5, "min_rank": 2,
         "max_single_ratio": 1.00, "cpc_factor": 1.00, "color": "FFF2CC"},
    ]
    _progress("예산 최적화 시뮬레이션 중...", 0.60)
    print(f"  [시뮬레이터] 예산 {budget:,}원 최적화...")
    recommended, total_cost, standby_rows, all_options_map = \
        attach_budget_plan(recommended, budget,
                           brand_profile={"keyword_categories": _sim_cat_config})

    # 확장 시나리오
    _progress("확장 시나리오 계산 중...", 0.82)
    current_rows = [r for r in recommended if not r.get("not_selected", False)]
    not_sel_rows = [r for r in recommended if r.get("not_selected", False)]
    not_sel_kw_rows = [{
        "keyword":             row["keyword"],
        "category":            row.get("category", ""),
        "keyword_type":        row.get("keyword_type", "GENERIC"),
        "competition":         row.get("competition", "MID"),
        "monthlySearchVolume": row.get("monthlySearchVolume", 0),
        "topOfPageBid":        row.get("topOfPageBid", 1000),
        "pc_impr":             row.get("pc_impr", 0),
        "mo_impr":             row.get("mo_impr", 0),
    } for row in not_sel_rows]

    scenario_data = simulate_scenarios(current_rows, not_sel_kw_rows,
                                       budget, all_options_map)

    try:
        optimal_data = simulate_unconstrained_optimal(current_rows, not_sel_rows)
    except Exception as _e:
        print(f"  [시뮬레이터] 최적효율 계산 실패 (무시): {_e}")
        optimal_data = None

    _progress("완료", 1.0)

    summary_profile = {
        "brand_name": client_name,
        "category": "",
        "monthly_budget": budget,
        "campaign_goal": "예산 최적화",
    }
    summary = build_summary_text(summary_profile, rows, recommended)

    return {
        "brand_name":      client_name,
        "brand_category":  "예산 시뮬레이터",
        "monthly_budget":  budget,
        "rows":            rows,
        "recommended":     recommended,
        "category_desc":   {},
        "summary":         summary,
        "standby_rows":    standby_rows,
        "scenario_data":   scenario_data,
        "optimal_data":    optimal_data,
        "total_cost":      total_cost,
    }


def main():
    parser = argparse.ArgumentParser(description="멀티 브랜드 SA 키워드 제안서 생성")
    parser.add_argument("--profile", "-p", default="data/client_profile.json",
                        help="프로파일 JSON 경로")
    parser.add_argument("--brand", "-b", default=None,
                        help="특정 브랜드만 실행 (brand_key 값)")
    args = parser.parse_args()

    profile_data = load_multi_profile(args.profile)
    client_name  = profile_data.get("client", "광고주")

    # 멀티 브랜드 vs 싱글 브랜드 판별
    brands = profile_data.get("brands", [])
    if not brands:
        # 기존 단일 브랜드 형식 호환
        brands = [profile_data]

    # 특정 브랜드만 실행
    if args.brand:
        brands = [b for b in brands if b.get("brand_key") == args.brand]
        if not brands:
            print(f"❌ brand_key '{args.brand}' 를 찾을 수 없습니다.")
            return

    print(f"광고주: {client_name}")
    print(f"브랜드 수: {len(brands)}개")

    # 브랜드별 데이터 수집
    brand_results = []
    for brand_cfg in brands:
        brand_profile = make_brand_profile(profile_data, brand_cfg)
        brand_name    = brand_cfg.get("brand_name", "브랜드")
        result = run_single_brand(brand_profile, brand_name)
        brand_results.append(result)

    # 멀티 브랜드 엑셀 저장
    print(f"\n{'='*50}")
    print("엑셀 저장 중...")
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename  = f"output/{client_name}_proposal_{timestamp}.xlsx"
    os.makedirs("output", exist_ok=True)

    save_multi_brand_excel(brand_results, filename, client_name)
    print(f"✅ 저장 완료: {filename}")


def _write_overview_sheet(ws, brand_results, client_name):
    """Overview 시트 작성 - 전체 브랜드 요약."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    def _bd():
        s = Side(style="thin")
        return Border(left=s, right=s, top=s, bottom=s)

    def _cell(row, col, value, bold=False, bg=None, fg="000000", align_h="center", font_size=10):
        c = ws.cell(row=row, column=col, value=value)
        c.font = Font(bold=bold, size=font_size, color=fg)
        c.alignment = Alignment(horizontal=align_h, vertical="center")
        c.border = _bd()
        if bg:
            c.fill = PatternFill("solid", fgColor=bg)
        return c

    # 컬럼 너비
    col_widths = {1:3, 2:25, 3:18, 4:12, 5:12, 6:12, 7:10, 8:14}
    for c, w in col_widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w

    row = 1
    # 타이틀
    ws.merge_cells(f"B{row}:H{row}")
    c = ws.cell(row=row, column=2, value=f"■ {client_name} 검색광고 캠페인 운영 제안")
    c.font = Font(bold=True, size=14, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="1F4E79")
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border = _bd()
    ws.row_dimensions[row].height = 28
    row += 2

    # ── 섹션1: 예산 요약 ─────────────────────────────
    ws.merge_cells(f"B{row}:H{row}")
    c = ws.cell(row=row, column=2, value="1. 캠페인 운영 예산 요약")
    c.font = Font(bold=True, size=11)
    ws.row_dimensions[row].height = 16
    row += 1

    headers = ["브랜드", "카테고리", "월 예산", "예상 노출", "예상 클릭", "CTR", "예상 비용"]
    for i, h in enumerate(headers):
        _cell(row, 2+i, h, bold=True, bg="2E74B5", fg="FFFFFF")
    ws.row_dimensions[row].height = 15
    row += 1

    total_imp = total_clk = total_cost = 0
    for result in brand_results:
        brand = result["brand_name"]
        recs  = result["recommended"]
        imp = clk = cost = 0
        for r in recs:
            if not r.get("not_selected") and not r.get("is_fallback"):
                imp  += (r.get("pc_sim_impressions") or 0) + (r.get("mo_sim_impressions") or 0)
                clk  += (r.get("pc_sim_clicks") or 0) + (r.get("mo_sim_clicks") or 0)
                cost += (r.get("pc_sim_cost") or 0) + (r.get("mo_sim_cost") or 0)
        budget  = result.get("monthly_budget", 500000)
        cat_str = result.get("brand_category", "")
        ctr = round(clk/imp, 4) if imp > 0 else 0
        vals = [brand, cat_str, f"{budget:,}원", f"{imp:,}", f"{clk:,}",
                f"{ctr*100:.2f}%", f"{cost:,}원"]
        for i, v in enumerate(vals):
            _cell(row, 2+i, v, align_h="left" if i == 0 else "center")
        total_imp  += imp
        total_clk  += clk
        total_cost += cost
        ws.row_dimensions[row].height = 15
        row += 1

    total_ctr = round(total_clk/total_imp, 4) if total_imp > 0 else 0
    total_vals = ["합계", "", "", f"{total_imp:,}", f"{total_clk:,}",
                  f"{total_ctr*100:.2f}%", f"{total_cost:,}원"]
    for i, v in enumerate(total_vals):
        _cell(row, 2+i, v, bold=True, bg="D9E1F2", align_h="left" if i == 0 else "center")
    ws.row_dimensions[row].height = 15
    row += 2

    # ── 섹션2: 키워드 현황 ────────────────────────────
    ws.merge_cells(f"B{row}:H{row}")
    c = ws.cell(row=row, column=2, value="2. 브랜드별 키워드 현황")
    c.font = Font(bold=True, size=11)
    ws.row_dimensions[row].height = 16
    row += 1

    headers2 = ["브랜드", "전체 키워드", "Estimate 성공", "Fallback", "브랜드KW", "일반KW", "경쟁사KW"]
    for i, h in enumerate(headers2):
        _cell(row, 2+i, h, bold=True, bg="2E74B5", fg="FFFFFF")
    ws.row_dimensions[row].height = 15
    row += 1

    for result in brand_results:
        brand = result["brand_name"]
        recs  = result["recommended"]
        total = len([r for r in recs if not r.get("not_selected")])
        est   = len([r for r in recs if not r.get("not_selected") and not r.get("is_fallback")])
        fb    = total - est
        cat_cnt = {}
        for r in recs:
            if not r.get("not_selected"):
                c = r.get("category","")
                cat_cnt[c] = cat_cnt.get(c,0) + 1
        brand_kw   = sum(v for k,v in cat_cnt.items() if "브랜드" in k)
        general_kw = sum(v for k,v in cat_cnt.items() if "일반" in k)
        comp_kw    = sum(v for k,v in cat_cnt.items() if "경쟁사" in k)
        vals = [brand, total, est, fb, brand_kw, general_kw, comp_kw]
        for i, v in enumerate(vals):
            _cell(row, 2+i, v, align_h="left" if i == 0 else "center")
        ws.row_dimensions[row].height = 15
        row += 1


def save_multi_brand_excel(brand_results, filename, client_name, return_bytes=False):
    """
    멀티 브랜드 엑셀 저장.
    return_bytes=True 이면 파일 쓰기 없이 bytes 반환 (Streamlit Cloud 호환).
    """
    import io
    import openpyxl
    from modules.excel_writer import (
        _write_proposal_sheet, _write_expanded_sheet,
        _write_summary_sheet, _write_all_sheet, _write_standby_sheet,
        _write_optimal_sheet,
    )

    wb = openpyxl.Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    ws_overview = wb.create_sheet(title="Overview")
    _write_overview_sheet(ws_overview, brand_results, client_name)

    for result in brand_results:
        brand_name  = result["brand_name"]
        _safe = brand_name.translate(str.maketrans('', '', r'\/:?*[]'))
        short_name  = _safe[:12] if len(_safe) > 12 else _safe
        recommended = result["recommended"]

        rec_proposal = [r for r in recommended if not r.get("not_selected", False)]
        ws_proposal = wb.create_sheet(title=f"{short_name}_제안서")
        _write_proposal_sheet(
            ws_proposal,
            rec_proposal,
            brand_name,
            result["category_desc"],
        )

        ws_expanded = wb.create_sheet(title=f"{short_name}_확장제안")
        _write_expanded_sheet(
            ws_expanded,
            result["scenario_data"],
            brand_name,
            recommended,
        )

        optimal_data = result.get("optimal_data")
        if optimal_data and optimal_data.get("optimal_rows"):
            ws_optimal = wb.create_sheet(title=f"{short_name}_최적효율제안")
            _write_optimal_sheet(ws_optimal, optimal_data, brand_name)

    if return_bytes:
        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)
        return bio.read()

    wb.save(filename)


if __name__ == "__main__":
    main()