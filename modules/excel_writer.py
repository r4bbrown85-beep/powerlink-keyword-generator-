import os
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── 색상 ───────────────────────────────────────────────────────
COLOR_TITLE_BG    = "1F4E79"
COLOR_TITLE_FG    = "FFFFFF"
COLOR_CAT_BRAND   = "BDD7EE"
COLOR_CAT_PRODUCT = "C6EFCE"
COLOR_CAT_GENERAL = "FFF2CC"
COLOR_CAT_COMP    = "FCE4D6"
COLOR_HEADER_BG   = "D9E2F3"
COLOR_SUBTOTAL_BG = "F2F2F2"
COLOR_FALLBACK_BG = "FFF9C4"


def _cat_color(category, cat_config_map=None):
    # cat_config_map: {name: color_hex} — AI 생성 카테고리 색상 우선 적용
    if cat_config_map and category in cat_config_map:
        return cat_config_map[category]
    s = str(category)
    if "브랜드" in s:                     return COLOR_CAT_BRAND
    if "상품" in s or "제품" in s or "서비스" in s: return COLOR_CAT_PRODUCT
    if "경쟁사" in s or "competitor" in s.lower(): return COLOR_CAT_COMP
    return COLOR_CAT_GENERAL


def _cat_sort_key(name):
    """카테고리명을 브랜드→상품/제품→일반→경쟁사 순으로 정렬하는 키 함수."""
    n = str(name)
    if "브랜드" in n:                                 return 0
    if "상품" in n or "제품" in n or "서비스" in n:    return 1
    if "경쟁사" in n or "competitor" in n.lower():    return 3
    return 2


def _sorted_cats(cat_groups):
    """cat_groups dict의 키를 카테고리 순서대로 반환."""
    return sorted(cat_groups.keys(), key=_cat_sort_key)


def _bd():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)


def _bd_top_medium():
    """위쪽만 medium 테두리, 나머지 thin."""
    return Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="medium"), bottom=Side(style="thin"),
    )


def _prop_fill_row(ws, row_num, last_col, bg=None, height=16):
    """last_col까지 빈 셀을 채워 테두리 일관성 확보."""
    from openpyxl.cell.cell import MergedCell
    for c in range(2, last_col + 1):
        cell = ws.cell(row=row_num, column=c)
        if isinstance(cell, MergedCell):
            continue
        if cell.value is None:
            cell.value = ""
        cell.border = _bd()
        if bg:
            cell.fill = _fill(bg)
    ws.row_dimensions[row_num].height = height


def _fill_row(ws, row, col_start, col_end, bg=None, height=None):
    """지정 범위 셀에 테두리+배경 채우기 (병합 셀 건드리지 않음)."""
    from openpyxl.cell.cell import MergedCell
    for col in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=col)
        if not isinstance(cell, MergedCell):
            cell.border = _bd()
            if bg:
                cell.fill = _fill(bg)
    if height:
        ws.row_dimensions[row].height = height


def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def _safe_int(v):
    try:
        return int(float(v)) if v not in ("", None) else 0
    except Exception:
        return 0


def _safe_float(v):
    try:
        return float(v) if v not in ("", None) else 0.0
    except Exception:
        return 0.0


def _write_cell(ws, row, col, value, bold=False, bg=None,
                align_h="center", num_fmt=None, font_size=10,
                italic=False, font_color="000000"):
    """병합된 셀이면 건드리지 않는 안전한 쓰기 헬퍼."""
    from openpyxl.cell.cell import MergedCell
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        return cell
    cell.value     = value
    cell.font      = Font(bold=bold, size=font_size,
                          color=font_color, italic=italic)
    cell.alignment = Alignment(horizontal=align_h, vertical="center")
    cell.border    = _bd()
    if bg:
        cell.fill = _fill(bg)
    if num_fmt:
        cell.number_format = num_fmt
    return cell


# ── 메인 ──────────────────────────────────────────────────────
def save_proposal_excel(rows, recommended_rows, category_desc,
                        summary_text, advertiser, standby_rows=None,
                        scenario_data=None, optimal_data=None):
    os.makedirs("output", exist_ok=True)
    ts          = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = f"output/{advertiser}_proposal_keywords_{ts}.xlsx"

    # 예산외(not_selected) 키워드는 제안서 본문에서 제외
    # → 전체키워드 시트에는 남아있으므로 필요시 참고 가능
    rec_proposal = [r for r in recommended_rows if not r.get("not_selected", False)]
    rec_sorted = sorted(
        rec_proposal,
        key=lambda x: (
            _cat_sort_key(x.get("category", "일반 키워드")),
            -_safe_int(x.get("sim_cost", 0))
        )
    )

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    _write_proposal_sheet(wb.create_sheet("제안서"),    rec_sorted, advertiser, category_desc)
    _write_summary_sheet( wb.create_sheet("요약"),      rec_sorted, advertiser)
    _write_all_sheet(     wb.create_sheet("전체키워드"), rows)

    # 등록 대기 키워드 시트 추가 (검색량 없음)
    if standby_rows:
        _write_standby_sheet(wb.create_sheet("등록대기"), standby_rows, advertiser)

    # 확장 제안 시트 추가
    if scenario_data:
        _write_expanded_sheet(wb.create_sheet("확장제안"), scenario_data, advertiser, recommended_rows)

    # 최적 효율 제안 시트 추가 (예산 무제한)
    if optimal_data and optimal_data.get("optimal_rows"):
        _write_optimal_sheet(wb.create_sheet("최적효율제안"), optimal_data, advertiser)

    wb.save(output_path)
    print(f"✅ 엑셀 저장 완료: {output_path}")


# ── 제안서 시트 ───────────────────────────────────────────────
# 컬럼 레이아웃 (A=여백, B~N=데이터):
#  B:키워드  C:구분  D:PC입찰가  E:PC노출  F:PC클릭  G:PC비용  H:PC순위
#  I:MO입찰가  J:MO노출  K:MO클릭  L:MO비용  M:MO순위  N:비고
_PROP_LAST_COL = 14  # N

def _write_proposal_sheet(ws, rec_sorted, advertiser, category_desc, sa_memo=None,
                           pc_budget=None, mo_budget=None):
    # ── 열 너비 ──────────────────────────────────────────────────
    widths = {1:2, 2:22, 3:11, 4:9, 5:9, 6:8, 7:10, 8:6, 9:9, 10:9, 11:8, 12:10, 13:6, 14:10}
    for c, w in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w

    LAST = _PROP_LAST_COL
    LAST_L = get_column_letter(LAST)

    # ── 인쇄 설정 (가로 A4 맞춤) ──────────────────────────────
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    # ── 내부 헬퍼 ────────────────────────────────────────────────
    BG_TITLE   = "1F4E79"   # 진네이비
    BG_SEC     = "2E75B6"   # 섹션 헤더 파랑
    BG_SEC_SUB = "EBF3FB"   # 섹션 서브 연파랑
    BG_HDR     = "D6E4F7"   # 칼럼 헤더 하늘
    BG_SUB     = "E2EFDA"   # 소계 연초록
    BG_SEP     = "F7F9FC"   # 빈 구분행 거의흰색
    BG_ODD     = "FAFCFF"   # 데이터 홀수행 아주연한파랑

    def _sec_row(row_num, text, height=20):
        ws.merge_cells(f"B{row_num}:{LAST_L}{row_num}")
        _write_cell(ws, row_num, 2, text,
                    bold=True, bg=BG_SEC, font_color="FFFFFF", align_h="left", font_size=10)
        _prop_fill_row(ws, row_num, LAST, bg=BG_SEC, height=height)

    def _gap(row_num, height=6):
        _prop_fill_row(ws, row_num, LAST, bg=BG_SEP, height=height)

    row = 1

    # ── 타이틀 ────────────────────────────────────────────────────
    ws.merge_cells(f"B{row}:{LAST_L}{row}")
    _write_cell(ws, row, 2,
                f"  {advertiser}  검색광고 캠페인 운영 제안",
                bold=True, bg=BG_TITLE, font_color="FFFFFF", font_size=13, align_h="left")
    _prop_fill_row(ws, row, LAST, bg=BG_TITLE, height=30)
    row += 1
    _gap(row, 4); row += 1

    # ── AI 전략 인사이트 ─────────────────────────────────────────
    if sa_memo and sa_memo.strip():
        _sec_row(row, "◆  AI 전략 인사이트 (자동 생성)", 18); row += 1
        for line in sa_memo.strip().splitlines():
            if not line.strip():
                continue
            ws.merge_cells(f"B{row}:{LAST_L}{row}")
            _write_cell(ws, row, 2, f"  {line.strip()}",
                        font_size=9, font_color="1E3A5F", align_h="left", bg=BG_SEC_SUB)
            _prop_fill_row(ws, row, LAST, bg=BG_SEC_SUB, height=13)
            row += 1
        _gap(row); row += 1

    # ── 광고 그룹 전략 설명 ──────────────────────────────────────
    if isinstance(category_desc, dict) and category_desc:
        _sec_row(row, "◆  광고 그룹 전략", 18); row += 1
        for i, (cat, desc) in enumerate(category_desc.items(), 1):
            ws.merge_cells(f"B{row}:{LAST_L}{row}")
            _write_cell(ws, row, 2, f"  {i}. {cat}  —  {desc}",
                        font_size=9, font_color="1E3A5F", align_h="left", bg=BG_SEC_SUB)
            _prop_fill_row(ws, row, LAST, bg=BG_SEC_SUB, height=13)
            row += 1
        _gap(row); row += 1

    # ── 1) 예상 성과 요약 ─────────────────────────────────────────
    _sec_row(row, "1)  예상 성과 요약  (VAT 미포함)"); row += 1

    for i, h in enumerate(["구분","평균 입찰가","노출수","클릭수","CTR","비용(원)","평균 순위"]):
        _write_cell(ws, row, 2+i, h, bold=True, bg=BG_HDR, font_size=10)
    _prop_fill_row(ws, row, LAST, bg=BG_HDR, height=17)
    row += 1

    def _agg(lst): return int(sum(lst)/len(lst)) if lst else 0
    def _ctr(c, i): return round(c/i, 4) if i > 0 else 0.0

    pc_bids, pc_imp, pc_clk, pc_cost, pc_rnk = [], [], [], [], []
    mo_bids, mo_imp, mo_clk, mo_cost, mo_rnk = [], [], [], [], []
    for r in rec_sorted:
        pb_pc = _safe_int(r.get("proposed_bid_pc", 0) or r.get("proposed_bid", 0))
        pb_mo = _safe_int(r.get("proposed_bid_mo", 0) or r.get("proposed_bid", 0))
        if _safe_int(r.get("pc_sim_impressions", 0)) > 0:
            pc_bids.append(pb_pc); pc_imp.append(_safe_int(r.get("pc_sim_impressions", 0)))
            pc_clk.append(_safe_int(r.get("pc_sim_clicks", 0)))
            pc_cost.append(_safe_int(r.get("pc_sim_cost", 0)))
            pc_rnk.append(_safe_int(r.get("proposed_rank_pc", 0)))
        if _safe_int(r.get("mo_sim_impressions", 0)) > 0:
            mo_bids.append(pb_mo); mo_imp.append(_safe_int(r.get("mo_sim_impressions", 0)))
            mo_clk.append(_safe_int(r.get("mo_sim_clicks", 0)))
            mo_cost.append(_safe_int(r.get("mo_sim_cost", 0)))
            mo_rnk.append(_safe_int(r.get("proposed_rank_mo", 0)))

    _bgt_lbl = {
        "TOTAL": f"총 {(pc_budget or 0)+(mo_budget or 0):,}원" if pc_budget and mo_budget else "",
        "PC":    f"예산 {pc_budget:,}원" if pc_budget else "",
        "MO":    f"예산 {mo_budget:,}원" if mo_budget else "",
    }
    for lbl, bids, imp, clk, cst, rnk in [
        ("TOTAL", pc_bids+mo_bids, pc_imp+mo_imp, pc_clk+mo_clk, pc_cost+mo_cost, pc_rnk+mo_rnk),
        ("PC",    pc_bids, pc_imp, pc_clk, pc_cost, pc_rnk),
        ("MO",    mo_bids, mo_imp, mo_clk, mo_cost, mo_rnk),
    ]:
        _ls = f"{lbl}   {_bgt_lbl.get(lbl,'')}" if _bgt_lbl.get(lbl) else lbl
        _bg = "E8F4FD" if lbl == "TOTAL" else BG_SEP
        _write_cell(ws, row, 2, _ls, bold=(lbl == "TOTAL"), bg=_bg, align_h="left")
        _write_cell(ws, row, 3, _agg(bids),             bg=_bg, num_fmt="#,##0")
        _write_cell(ws, row, 4, sum(imp),               bg=_bg, num_fmt="#,##0")
        _write_cell(ws, row, 5, sum(clk),               bg=_bg, num_fmt="#,##0")
        _write_cell(ws, row, 6, _ctr(sum(clk),sum(imp)), bg=_bg, num_fmt="0.00%")
        _write_cell(ws, row, 7, sum(cst),               bg=_bg, num_fmt="#,##0")
        _write_cell(ws, row, 8, _agg(rnk),              bg=_bg, num_fmt="0.0")
        _prop_fill_row(ws, row, LAST, bg=_bg, height=16)
        row += 1
    _gap(row); row += 1

    # ── 2) 키워드별 상세 예상 성과 ───────────────────────────────
    _sec_row(row, "2)  키워드별 상세 예상 성과  (VAT 미포함)"); row += 1

    # ── 카테고리별 키워드 출력 ───────────────────────────────────
    cat_groups = {}
    for r in rec_sorted:
        cat_groups.setdefault(r.get("category", "일반 키워드"), []).append(r)

    for cat_num, cat in enumerate(_sorted_cats(cat_groups), 1):
        kws = cat_groups[cat]
        cc  = _cat_color(cat)

        # 카테고리 소제목 (medium 상단 테두리)
        ws.merge_cells(f"B{row}:{LAST_L}{row}")
        cell = ws.cell(row=row, column=2)
        cell.value     = f"  {chr(9312 + cat_num - 1)}  {cat}"
        cell.font      = Font(bold=True, size=10, color="000000")
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.fill      = _fill(cc)
        cell.border    = _bd_top_medium()
        _prop_fill_row(ws, row, LAST, bg=cc, height=18)
        # medium 테두리 일관성: 병합 영역 왼쪽 셀에만 적용됐으므로 OK
        row += 1

        # 칼럼 헤더 1행 (PC / MO 그룹)
        _write_cell(ws, row, 2, "키워드",            bold=True, bg=BG_HDR, font_size=10)
        _write_cell(ws, row, 3, "구분",              bold=True, bg=BG_HDR, font_size=10)
        ws.merge_cells(f"D{row}:H{row}")
        _write_cell(ws, row, 4, "PC 예상 운영 성과", bold=True, bg="BBDEFB", align_h="center", font_size=10)
        ws.merge_cells(f"I{row}:M{row}")
        _write_cell(ws, row, 9, "MO 예상 운영 성과", bold=True, bg="C8E6C9", align_h="center", font_size=10)
        _write_cell(ws, row, 14, "비고",             bold=True, bg=BG_HDR, font_size=10)
        _prop_fill_row(ws, row, LAST, bg=BG_HDR, height=16)
        row += 1

        # 칼럼 헤더 2행 (세부 항목)
        _write_cell(ws, row, 2, "",  bold=True, bg=BG_HDR)
        _write_cell(ws, row, 3, "",  bold=True, bg=BG_HDR)
        for _ci, _lbl in enumerate(["입찰가","노출수","클릭수","비용(원)","순위"]):
            _write_cell(ws, row, 4+_ci, _lbl, bold=True, bg="BBDEFB", font_size=9)
        for _ci, _lbl in enumerate(["입찰가","노출수","클릭수","비용(원)","순위"]):
            _write_cell(ws, row, 9+_ci, _lbl, bold=True, bg="C8E6C9", font_size=9)
        _write_cell(ws, row, 14, "", bold=True, bg=BG_HDR)
        _prop_fill_row(ws, row, LAST, bg=BG_HDR, height=15)
        row += 1

        for idx, kw_row in enumerate(kws):
            fb           = kw_row.get("is_fallback", False)
            not_selected = kw_row.get("not_selected", False)
            if fb:
                row_bg = COLOR_FALLBACK_BG
            elif not_selected:
                row_bg = "EEEEEE"
            else:
                row_bg = BG_ODD if idx % 2 == 1 else None

            pc_bid   = _safe_int(kw_row.get("proposed_bid_pc", 0) or kw_row.get("proposed_bid", 0))
            mo_bid   = _safe_int(kw_row.get("proposed_bid_mo", 0) or kw_row.get("proposed_bid", 0))
            pc_impr  = _safe_int(kw_row.get("pc_sim_impressions", 0))
            pc_clks  = _safe_int(kw_row.get("pc_sim_clicks", 0))
            pc_cst   = _safe_int(kw_row.get("pc_sim_cost", 0))
            pc_rnk   = _safe_int(kw_row.get("proposed_rank_pc", 0))
            mo_impr  = _safe_int(kw_row.get("mo_sim_impressions", 0))
            mo_clks  = _safe_int(kw_row.get("mo_sim_clicks", 0))
            mo_cst   = _safe_int(kw_row.get("mo_sim_cost", 0))
            mo_rnk   = _safe_int(kw_row.get("proposed_rank_mo", 0))

            has_perf = not fb and not not_selected

            def _v(val, is_rank=False):
                if has_perf:
                    return val or ""
                if fb and not not_selected and is_rank:
                    return val or ""
                return ""

            # 비고
            already_reg = kw_row.get("already_registered", False)
            actual_rank = kw_row.get("actual_avg_rank", 0)
            if fb:
                _fb_r = kw_row.get("proposed_rank_pc") or kw_row.get("proposed_rank", "")
                note = f"추정·{_fb_r}위" if _fb_r else "추정"
            elif not_selected:
                note = "예산외"
            elif already_reg and actual_rank > 0:
                note = f"기등록·실측{actual_rank:.0f}위"
            elif already_reg:
                note = "기등록"
            elif actual_rank > 0:
                note = f"실측{actual_rank:.0f}위"
            else:
                note = ""
            dev_eff = kw_row.get("device_efficiency", "")
            if dev_eff:
                tag = {"PC우위": "PC↑", "MO우위": "MO↑", "균등": "균등"}.get(dev_eff, "")
                if tag:
                    note = f"{note}  {tag}" if note else tag

            _write_cell(ws, row,  2, kw_row.get("keyword", ""),    bg=row_bg, align_h="left", font_size=10)
            _write_cell(ws, row,  3, kw_row.get("category", ""),   bg=row_bg, font_size=9)
            _write_cell(ws, row,  4, pc_bid or "",                  bg=row_bg, num_fmt="#,##0")
            _write_cell(ws, row,  5, _v(pc_impr),                   bg=row_bg, num_fmt="#,##0")
            _write_cell(ws, row,  6, _v(pc_clks),                   bg=row_bg, num_fmt="#,##0")
            _write_cell(ws, row,  7, _v(pc_cst),                    bg=row_bg, num_fmt="#,##0")
            _write_cell(ws, row,  8, _v(pc_rnk, is_rank=True),      bg=row_bg)
            _write_cell(ws, row,  9, mo_bid or "",                  bg=row_bg, num_fmt="#,##0")
            _write_cell(ws, row, 10, _v(mo_impr),                   bg=row_bg, num_fmt="#,##0")
            _write_cell(ws, row, 11, _v(mo_clks),                   bg=row_bg, num_fmt="#,##0")
            _write_cell(ws, row, 12, _v(mo_cst),                    bg=row_bg, num_fmt="#,##0")
            _write_cell(ws, row, 13, _v(mo_rnk, is_rank=True),      bg=row_bg)
            _write_cell(ws, row, 14, note,                          bg=row_bg, font_size=9)
            _prop_fill_row(ws, row, LAST, bg=row_bg, height=16)
            row += 1

        # 카테고리 소계
        c_pc_imp  = sum(_safe_int(r.get("pc_sim_impressions",0)) for r in kws)
        c_pc_clk  = sum(_safe_int(r.get("pc_sim_clicks",0))      for r in kws)
        c_pc_cost = sum(_safe_int(r.get("pc_sim_cost",0))        for r in kws)
        c_mo_imp  = sum(_safe_int(r.get("mo_sim_impressions",0)) for r in kws)
        c_mo_clk  = sum(_safe_int(r.get("mo_sim_clicks",0))      for r in kws)
        c_mo_cost = sum(_safe_int(r.get("mo_sim_cost",0))        for r in kws)

        ws.merge_cells(f"B{row}:C{row}")
        _write_cell(ws, row,  2, f"  소계 ({len(kws)}개)", bold=True, bg=BG_SUB, align_h="left")
        _write_cell(ws, row,  4, "",        bg=BG_SUB)
        _write_cell(ws, row,  5, c_pc_imp,  bg=BG_SUB, num_fmt="#,##0", bold=True)
        _write_cell(ws, row,  6, c_pc_clk,  bg=BG_SUB, num_fmt="#,##0", bold=True)
        _write_cell(ws, row,  7, c_pc_cost, bg=BG_SUB, num_fmt="#,##0", bold=True)
        _write_cell(ws, row,  8, "",        bg=BG_SUB)
        _write_cell(ws, row,  9, "",        bg=BG_SUB)
        _write_cell(ws, row, 10, c_mo_imp,  bg=BG_SUB, num_fmt="#,##0", bold=True)
        _write_cell(ws, row, 11, c_mo_clk,  bg=BG_SUB, num_fmt="#,##0", bold=True)
        _write_cell(ws, row, 12, c_mo_cost, bg=BG_SUB, num_fmt="#,##0", bold=True)
        _write_cell(ws, row, 13, "",        bg=BG_SUB)
        _write_cell(ws, row, 14, "",        bg=BG_SUB)
        _prop_fill_row(ws, row, LAST, bg=BG_SUB, height=16)
        row += 1
        _gap(row, 5); row += 1

    # ── 예산 합계 ─────────────────────────────────────────────────
    total_pc = sum(_safe_int(r.get("pc_sim_cost", 0)) for r in rec_sorted)
    total_mo = sum(_safe_int(r.get("mo_sim_cost", 0)) for r in rec_sorted)

    for lbl, val, bg_c in [
        ("PC 캠페인 예상 집행 금액", total_pc,        "BBDEFB"),
        ("MO 캠페인 예상 집행 금액", total_mo,        "C8E6C9"),
        ("합계",                    total_pc+total_mo, BG_SUB),
    ]:
        ws.merge_cells(f"B{row}:D{row}")
        _write_cell(ws, row, 2, lbl,  bold=True, bg=bg_c, align_h="left")
        _write_cell(ws, row, 5, val,  bold=True, bg=bg_c, num_fmt="#,##0")
        _prop_fill_row(ws, row, LAST, bg=bg_c, height=16)
        row += 1

    _gap(row, 4); row += 1

    # ── 주석 ─────────────────────────────────────────────────────
    ws.merge_cells(f"B{row}:{LAST_L}{row}")
    _write_cell(ws, row, 2,
                "※ 입찰가·성과 수치는 네이버 성과예측 데이터 기반 추정값이며, "
                "실제 운영 결과와 다를 수 있습니다.  "
                "검색량이 적은 키워드(추정)는 동일 카테고리 평균 기준으로 제안합니다.",
                font_size=9, font_color="6B6B6B", italic=True, align_h="left", bg="FAFAFA")
    _prop_fill_row(ws, row, LAST, bg="FAFAFA", height=18)

    # ── 틀 고정 (헤더 위 2행) ────────────────────────────────────
    ws.freeze_panes = "B3"


# ── 요약 시트 ─────────────────────────────────────────────────
def _write_summary_sheet(ws, rec_sorted, advertiser, pc_budget=None, mo_budget=None):
    for col, w in {1:3, 2:26, 3:18, 4:18, 5:18}.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    row = 1
    ws.merge_cells(f"B{row}:E{row}")
    _write_cell(ws, row, 2, f"■ {advertiser} SA 광고 제안 요약",
                bold=True, font_size=13, bg=COLOR_TITLE_BG,
                font_color=COLOR_TITLE_FG, align_h="left")
    ws.row_dimensions[row].height = 22
    row += 2

    _write_cell(ws, row, 2, "생성일시", bold=True, bg=COLOR_HEADER_BG)
    ws.merge_cells(f"C{row}:E{row}")
    _write_cell(ws, row, 3, datetime.now().strftime("%Y-%m-%d %H:%M"), align_h="left")
    row += 2

    total_kw      = len(rec_sorted)
    total_impr    = sum(_safe_int(r.get("sim_impressions",0))    for r in rec_sorted)
    total_clk     = sum(_safe_int(r.get("sim_clicks",0))         for r in rec_sorted)
    total_cost    = sum(_safe_int(r.get("sim_cost",0))           for r in rec_sorted)
    total_pc_clk  = sum(_safe_int(r.get("pc_sim_clicks",0))      for r in rec_sorted)
    total_mo_clk  = sum(_safe_int(r.get("mo_sim_clicks",0))      for r in rec_sorted)
    total_pc_cost = sum(_safe_int(r.get("pc_sim_cost",0))        for r in rec_sorted)
    total_mo_cost = sum(_safe_int(r.get("mo_sim_cost",0))        for r in rec_sorted)
    avg_cpc       = int(total_cost / total_clk) if total_clk > 0 else 0
    avg_ctr       = round(total_clk / total_impr, 4) if total_impr > 0 else 0.0
    fb_cnt        = sum(1 for r in rec_sorted if r.get("is_fallback", False))

    # 실제 운영 예산 = pc_sim_cost + mo_sim_cost (PC/MO 분리 캠페인 기준)
    real_total_cost = total_pc_cost + total_mo_cost

    for label, value, fmt in [
        ("제안 키워드 수",   total_kw,        "#,##0"),
        ("추정값 키워드 수", fb_cnt,          "#,##0"),
        ("예상 총 노출수",   total_impr,      "#,##0"),
        ("예상 총 클릭수",   total_clk,       "#,##0"),
        ("  ├ PC 클릭수",    total_pc_clk,    "#,##0"),
        ("  └ MO 클릭수",    total_mo_clk,    "#,##0"),
        ("평균 CPC",         avg_cpc,         "#,##0"),
        ("평균 CTR",         avg_ctr,         "0.00%"),
    ]:
        is_sub = "  " in label
        _write_cell(ws, row, 2, label, bold=not is_sub,
                    bg=COLOR_HEADER_BG if not is_sub else None)
        _write_cell(ws, row, 3, value, num_fmt=fmt, align_h="right")
        row += 1

    row += 1
    # 운영 예산 섹션 (PC/MO 캠페인 분리 기준)
    ws.merge_cells(f"B{row}:E{row}")
    _write_cell(ws, row, 2, "■ 운영 예산 (PC/MO 캠페인 분리 기준)",
                bold=True, align_h="left")
    row += 1

    if pc_budget and mo_budget:
        # 디바이스별 예산 분리 모드: 입력 예산 + 예상 집행 + 활용률 표시
        for dev, bgt, actual, bg_col in [
            ("PC 예산", pc_budget, total_pc_cost, "E3F2FD"),
            ("MO 예산", mo_budget, total_mo_cost, "E8F5E9"),
        ]:
            util = round(actual / bgt * 100, 1) if bgt > 0 else 0
            ws.merge_cells(f"C{row}:D{row}")
            _write_cell(ws, row, 2, dev, bold=True, bg=COLOR_HEADER_BG)
            _write_cell(ws, row, 3, bgt,    num_fmt="#,##0", align_h="right", bg=bg_col, bold=True)
            _write_cell(ws, row, 4, actual, num_fmt="#,##0", align_h="right", bg=bg_col)
            _write_cell(ws, row, 5, f"{util}%",              align_h="right", bg=bg_col)
            row += 1
        # 헤더 표기 (거꾸로 삽입)
        for _c, _lbl in [(2,"구분"), (3,"입력 예산"), (4,"예상 집행"), (5,"활용률")]:
            _write_cell(ws, row-3, _c, _lbl, bold=True, bg=COLOR_HEADER_BG)  # no-op already written
        ws.merge_cells(f"C{row}:E{row}")
        _write_cell(ws, row, 2, "합계", bold=True, bg=COLOR_SUBTOTAL_BG)
        _write_cell(ws, row, 3, pc_budget + mo_budget, num_fmt="#,##0", align_h="right",
                    bg=COLOR_SUBTOTAL_BG, bold=True)
        row += 1
        ws.merge_cells(f"B{row}:E{row}")
        _write_cell(ws, row, 2,
                    f"※ PC {total_pc_cost:,}원 / MO {total_mo_cost:,}원 (입력 예산 PC {pc_budget:,} / MO {mo_budget:,})",
                    font_size=8, font_color="595959", italic=True, align_h="left")
        row += 1
    else:
        ws.merge_cells(f"C{row}:E{row}")
        _write_cell(ws, row, 2, "PC 캠페인 예산", bold=True, bg=COLOR_HEADER_BG)
        _write_cell(ws, row, 3, total_pc_cost, num_fmt="#,##0", align_h="right",
                    bg="E3F2FD", bold=True)
        row += 1

        ws.merge_cells(f"C{row}:E{row}")
        _write_cell(ws, row, 2, "MO 캠페인 예산", bold=True, bg=COLOR_HEADER_BG)
        _write_cell(ws, row, 3, total_mo_cost, num_fmt="#,##0", align_h="right",
                    bg="E8F5E9", bold=True)
        row += 1

        ws.merge_cells(f"C{row}:E{row}")
        _write_cell(ws, row, 2, "합계", bold=True, bg=COLOR_SUBTOTAL_BG)
        _write_cell(ws, row, 3, real_total_cost, num_fmt="#,##0", align_h="right",
                    bg=COLOR_SUBTOTAL_BG, bold=True)
        row += 1

        # 안내 문구
        ws.merge_cells(f"B{row}:E{row}")
        _write_cell(ws, row, 2,
                    f"※ PC/MO 캠페인 분리 운영 기준입니다. 총 예산 내에서 PC {total_pc_cost:,}원 / MO {total_mo_cost:,}원으로 배분하여 운영하세요.",
                    font_size=8, font_color="595959", italic=True, align_h="left")
        row += 1

    row += 1
    ws.merge_cells(f"B{row}:E{row}")
    _write_cell(ws, row, 2, "■ 카테고리별 예상 성과", bold=True, align_h="left")
    row += 1

    for i, h in enumerate(["카테고리","키워드수","예상클릭","예상비용"]):
        _write_cell(ws, row, 2+i, h, bold=True, bg=COLOR_HEADER_BG)
    row += 1

    cat_groups = {}
    for r in rec_sorted:
        cat_groups.setdefault(r.get("category","일반 키워드"), []).append(r)

    for cat in _sorted_cats(cat_groups):
        kws = cat_groups[cat]
        bg  = _cat_color(cat)
        _write_cell(ws, row, 2, cat, bg=bg)
        _write_cell(ws, row, 3, len(kws), bg=bg, num_fmt="#,##0")
        _write_cell(ws, row, 4, sum(_safe_int(r.get("sim_clicks",0)) for r in kws), bg=bg, num_fmt="#,##0")
        _write_cell(ws, row, 5, sum(_safe_int(r.get("sim_cost",0))   for r in kws), bg=bg, num_fmt="#,##0")
        row += 1


# ── 전체키워드 시트 ───────────────────────────────────────────
def _write_all_sheet(ws, rows):
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 14
    for i in range(3, 15):
        ws.column_dimensions[get_column_letter(i)].width = 13

    # pc_impr/mo_impr = 월간 검색량(쿼리수), 광고 노출수 아님
    # pc_click/mo_click = 시장 전체 평균 클릭수 (키워드툴 API 기준)
    headers = ["키워드","카테고리","키워드유형","기본점수","추천점수","소스",
               "PC월검색량","MO월검색량","PC평균클릭(시장)","MO평균클릭(시장)","CTR",
               "경쟁도","기준입찰가","월검색량합계"]
    for i, h in enumerate(headers, 1):
        _write_cell(ws, 1, i, h, bold=True, bg=COLOR_HEADER_BG)

    for idx, r in enumerate(rows, 2):
        values = [
            r.get("keyword",""),
            r.get("category",""),
            r.get("keyword_type",""),
            _safe_int(r.get("score",0)),
            _safe_int(r.get("recommendation_score",0)),
            r.get("source",""),
            _safe_int(r.get("pc_impr",0)),
            _safe_int(r.get("mo_impr",0)),
            _safe_float(r.get("pc_click",0)),
            _safe_float(r.get("mo_click",0)),
            _safe_float(r.get("ctr",0)),
            r.get("competition",""),
            _safe_int(r.get("topOfPageBid",0)),
            _safe_int(r.get("monthlySearchVolume",0)),
        ]
        for ci, val in enumerate(values, 1):
            _write_cell(ws, idx, ci, val,
                        align_h="left" if ci <= 2 else "center",
                        font_size=9)

# ── 등록 대기 키워드 시트 ──────────────────────────────────────────────────
def _write_standby_sheet(ws, standby_rows, advertiser):
    """
    검색량이 없어서 성과 추정이 불가능한 키워드.
    브랜드/제품 인지도가 올라가면 자동으로 노출되기 시작하는 키워드들.
    입찰가만 등록해두는 '씨앗 키워드' 개념.
    """
    for col, w in {1:3, 2:35, 3:18, 4:12}.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    row = 1

    # 타이틀
    ws.merge_cells(f"B{row}:D{row}")
    _write_cell(ws, row, 2,
                f"▪ {advertiser} 등록 대기 키워드",
                bold=True, bg=COLOR_TITLE_BG,
                font_color=COLOR_TITLE_FG, font_size=13, align_h="left")
    ws.row_dimensions[row].height = 22
    row += 1

    # 안내문
    ws.merge_cells(f"B{row}:D{row}")
    _write_cell(ws, row, 2,
                "현재 검색량이 없어 예상 성과 산출이 불가한 키워드입니다. ",
                font_size=9, font_color="595959", italic=True, align_h="left")
    row += 1
    ws.merge_cells(f"B{row}:D{row}")
    _write_cell(ws, row, 2,
                "브랜드/제품 인지도 상승 시 자연스럽게 검색량이 발생하므로 사전에 등록해두는 것을 권장합니다.",
                font_size=9, font_color="595959", italic=True, align_h="left")
    row += 2

    # 카테고리별 정렬
    cat_groups = {}
    for r in standby_rows:
        cat = r.get("category", "일반 키워드")
        cat_groups.setdefault(cat, []).append(r)

    total = 0
    for cat in _sorted_cats(cat_groups):
        kws = cat_groups[cat]
        cc  = _cat_color(cat)

        # 카테고리 소제목
        ws.merge_cells(f"B{row}:D{row}")
        _write_cell(ws, row, 2, cat,
                    bold=True, bg=cc, align_h="left")
        ws.row_dimensions[row].height = 16
        row += 1

        # 헤더
        for col, label in [(2, "키워드"), (3, "구분"), (4, "권장입찰가")]:
            _write_cell(ws, row, col, label, bold=True, bg=COLOR_HEADER_BG)
        row += 1

        for kw_row in sorted(kws, key=lambda x: x.get("keyword", "")):
            _write_cell(ws, row, 2, kw_row.get("keyword", ""), align_h="left")
            _write_cell(ws, row, 3, cat)
            _write_cell(ws, row, 4, 70, num_fmt="#,##0")  # 최소 입찰가 70원
            row += 1
            total += 1

        # 소계 (키워드 칼럼 비워서 오인 방지)
        _write_cell(ws, row, 2, "", bg=COLOR_SUBTOTAL_BG, align_h="left")
        _write_cell(ws, row, 3, f"◀ {cat} 소계 ({len(kws)}개)",
                    bold=True, bg=COLOR_SUBTOTAL_BG, align_h="left")
        _write_cell(ws, row, 4, "", bg=COLOR_SUBTOTAL_BG)
        row += 2

    # 합계
    ws.merge_cells(f"B{row}:D{row}")
    _write_cell(ws, row, 2,
                f"총 {total}개 키워드 | 검색량 발생 시 자동 노출 시작",
                bold=True, bg=COLOR_SUBTOTAL_BG, align_h="left")


# ── 확장 제안 시트 (개편) ──────────────────────────────────────────────────
def _write_expanded_sheet(ws, scenario_data, advertiser, recommended_rows):
    """
    3섹션 구성:
      섹션 1: 예산 시나리오 비교 (현재/+40%/+100%)
      섹션 2: 예산 추가시 진입 가능한 신규 키워드
      섹션 3: 현재 키워드 중 입찰가 올리면 좋은 것들
    """
    for col, w in {1:3, 2:32, 3:12, 4:10, 5:10, 6:9, 7:9, 8:9, 9:9, 10:8, 11:8, 12:10, 13:10}.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    bid_up_rows       = scenario_data.get("bid_up_rows", [])
    new_kw_rows       = scenario_data.get("new_kw_rows", [])
    scenarios         = scenario_data.get("scenarios", [])
    bid_up_status     = scenario_data.get("bid_up_status", "no_upgrade")
    all_kws_included  = scenario_data.get("all_kws_included", False)
    current_kw_count  = scenario_data.get("current_kw_count", 0)

    row = 1

    # ── 타이틀 ──
    ws.merge_cells(f"B{row}:J{row}")
    _write_cell(ws, row, 2, f"▪ {advertiser} 확장 제안",
                bold=True, bg=COLOR_TITLE_BG, font_color=COLOR_TITLE_FG,
                font_size=13, align_h="left")
    ws.row_dimensions[row].height = 22
    row += 1
    ws.merge_cells(f"B{row}:J{row}")
    _write_cell(ws, row, 2,
                "현재 예산 기준 제안과 예산 확대 시 추가 효과를 시나리오별로 비교합니다.",
                font_size=9, font_color="595959", italic=True, align_h="left")
    ws.row_dimensions[row].height = 14
    row += 2
    # freeze_panes는 시트 상단 고정 (타이틀 2행 + 공백 1행 아래부터 스크롤)
    ws.freeze_panes = f"B{row}"

    # ══════════════════════════════════════════════
    # 섹션 1: 예산 시나리오 비교
    # ══════════════════════════════════════════════
    ws.merge_cells(f"B{row}:J{row}")
    _write_cell(ws, row, 2, "▶ SECTION 1. 예산 시나리오 비교",
                bold=True, bg="1F4E79", font_color="FFFFFF", font_size=11, align_h="left")
    ws.row_dimensions[row].height = 20
    row += 1

    # 헤더
    for col, label in [(2,"시나리오"), (3,"예산"), (4,"총 키워드"), (5,"예상 클릭"), (6,"예상 비용"),
                        (7,"추가 키워드"), (8,"추가 클릭"), (9,"클릭 증가율")]:
        _write_cell(ws, row, col, label, bold=True, bg=COLOR_HEADER_BG)
    row += 1

    base_clicks = scenarios[0]["clicks"] if scenarios else 1
    all_zero_expansion = all(sc.get("add_clicks", 0) == 0 for sc in scenarios[1:]) if len(scenarios) > 1 else True
    for i, sc in enumerate(scenarios):
        bg = "E8F5E9" if i == 0 else ("FFF9C4" if i == 1 else "FFF3E0")
        if i == 0:
            click_rate = "기준"
        elif all_zero_expansion and all_kws_included:
            click_rate = "키워드 추가 필요"
        elif i > 1 and sc["cost"] == scenarios[i - 1]["cost"]:
            click_rate = "추가 집행 가능 키워드 없음"
        else:
            click_rate = f"+{(sc['clicks']/base_clicks - 1)*100:.0f}%"
        _write_cell(ws, row, 2, sc["label"],       bg=bg, align_h="left", bold=(i==0))
        _write_cell(ws, row, 3, sc["budget"],      bg=bg, num_fmt="#,##0")
        _write_cell(ws, row, 4, sc["kw_count"],    bg=bg, num_fmt="#,##0")
        _write_cell(ws, row, 5, sc["clicks"],      bg=bg, num_fmt="#,##0")
        _write_cell(ws, row, 6, sc["cost"],        bg=bg, num_fmt="#,##0")
        _write_cell(ws, row, 7, sc["add_kws"] if i > 0 else "-",    bg=bg)
        _write_cell(ws, row, 8, sc["add_clicks"] if i > 0 else "-", bg=bg, num_fmt="#,##0")
        _write_cell(ws, row, 9, click_rate,        bg=bg, bold=(i>0))
        ws.row_dimensions[row].height = 16
        row += 1
    # 예산 여유 있지만 추가 키워드 없는 경우 안내
    if all_zero_expansion and all_kws_included:
        ws.merge_cells(f"B{row}:I{row}")
        _write_cell(ws, row, 2,
                    f"ℹ️  현재 입력된 {current_kw_count}개 키워드가 모두 예산 내에서 운영 중입니다. "
                    "예산 확대 효과를 보려면 키워드를 추가로 입력하세요.",
                    font_color="1565C0", align_h="left", font_size=9, italic=True)
        ws.row_dimensions[row].height = 14
        row += 1
    row += 1

    # ══════════════════════════════════════════════
    # 섹션 2: 신규 진입 가능 키워드
    # ══════════════════════════════════════════════
    # 현재 제안 비용 (섹션1 첫번째 시나리오)
    cur_cost = scenarios[0]["cost"] if scenarios else 0
    ws.merge_cells(f"B{row}:J{row}")
    _write_cell(ws, row, 2, "▶ SECTION 2. 예산 확대 시 신규 진입 가능 키워드",
                bold=True, bg="1F4E79", font_color="FFFFFF", font_size=11, align_h="left")
    ws.row_dimensions[row].height = 20
    row += 1
    ws.merge_cells(f"B{row}:J{row}")
    _write_cell(ws, row, 2,
                "현재 예산 초과로 제외된 키워드입니다. 예산 추가 시 우선순위 순으로 진입을 권장합니다.",
                font_size=9, font_color="595959", italic=True, align_h="left")
    row += 1

    if not new_kw_rows:
        ws.merge_cells(f"B{row}:J{row}")
        if all_kws_included:
            _write_cell(ws, row, 2,
                        f"✅  입력된 {current_kw_count}개 키워드가 모두 현재 예산으로 운영되고 있습니다. "
                        "추가 키워드를 입력하면 이 섹션에 진입 후보가 표시됩니다.",
                        font_color="1F7F3F", align_h="left")
        else:
            _write_cell(ws, row, 2, "신규 진입 가능 키워드가 없습니다.", font_color="888888", align_h="left")
        ws.row_dimensions[row].height = 16
        row += 2
    else:
        # 헤더
        # 시나리오 라벨 추출 (2번째, 3번째 시나리오)
        sc_labels = [sc["label"] for sc in scenarios[1:3]] if len(scenarios) > 1 else []
        sc_budgets = [sc["budget"] for sc in scenarios[1:3]] if len(scenarios) > 1 else []

        for col, label in [(2,"키워드"), (3,"구분"), (4,"PC입찰가"), (5,"MO입찰가"),
                            (6,"PC클릭"), (7,"MO클릭"), (8,"PC비용"), (9,"MO비용"),
                            (10,"PC순위"), (11,"MO순위")]:
            _write_cell(ws, row, col, label, bold=True, bg=COLOR_HEADER_BG)
        # 시나리오별 포함여부 헤더
        for i, sc_label in enumerate(sc_labels):
            _write_cell(ws, row, 12+i, sc_label, bold=True, bg="FFF9C4")
        row += 1

        # 카테고리별로 묶어서 출력
        cat_groups = {}
        for r in new_kw_rows:
            cat = r.get("category", "일반 키워드")
            cat_groups.setdefault(cat, []).append(r)

        cumul_cost = 0
        cumul_kw   = 0
        # 시나리오별 누적 비용: 카테고리를 넘어 전체 누적으로 포함여부 판단
        sc_cumul_carry = [0] * len(sc_budgets)

        for cat in _sorted_cats(cat_groups):
            kws = cat_groups[cat]
            cc  = _cat_color(cat)
            ws.merge_cells(f"B{row}:J{row}")
            _write_cell(ws, row, 2, cat, bold=True, bg=cc, align_h="left")
            row += 1

            for kw_row in kws:
                kw_cost = kw_row.get("pc_cost", 0) + kw_row.get("mo_cost", 0)
                cumul_cost += kw_cost
                cumul_kw   += 1

                # 시나리오별 포함 여부 계산
                # 현재 예산에서 각 시나리오 추가 예산 내에 들어오는지
                sc_status = []
                for i, sc_budget in enumerate(sc_budgets):
                    add_budget = sc_budget - cur_cost  # 추가 가능 예산
                    sc_cumul_carry[i] += kw_cost
                    if sc_cumul_carry[i] <= add_budget:
                        sc_status.append(("포함", "E8F5E9"))   # 초록
                    else:
                        sc_status.append(("예산초과", "FFEBEE"))  # 빨강

                _write_cell(ws, row, 2, kw_row.get("keyword", ""),     align_h="left")
                _write_cell(ws, row, 3, cat)
                _write_cell(ws, row, 4, kw_row.get("pc_bid", 0),       num_fmt="#,##0")
                _write_cell(ws, row, 5, kw_row.get("mo_bid", 0),       num_fmt="#,##0")
                _write_cell(ws, row, 6, kw_row.get("pc_clicks", 0),    num_fmt="#,##0")
                _write_cell(ws, row, 7, kw_row.get("mo_clicks", 0),    num_fmt="#,##0")
                _write_cell(ws, row, 8, kw_row.get("pc_cost", 0),      num_fmt="#,##0")
                _write_cell(ws, row, 9, kw_row.get("mo_cost", 0),      num_fmt="#,##0")
                _write_cell(ws, row, 10, kw_row.get("rank_pc", "-"))
                _write_cell(ws, row, 11, kw_row.get("rank_mo", "-"))
                for i, (status, bg) in enumerate(sc_status):
                    _write_cell(ws, row, 12 + i, status, bg=bg, bold=(status == "포함"))
                row += 1

        # 소계
        _write_cell(ws, row, 2, f"신규 진입 키워드 합계: {cumul_kw}개",
                    bold=True, bg=COLOR_SUBTOTAL_BG, align_h="left")
        ws.merge_cells(f"C{row}:G{row}")
        _write_cell(ws, row, 3, f"추가 예상 비용: {cumul_cost:,}원",
                    bold=True, bg=COLOR_SUBTOTAL_BG, align_h="left")
        row += 2

    # ══════════════════════════════════════════════
    # 섹션 3: 입찰가 올리면 좋은 키워드
    # ══════════════════════════════════════════════
    ws.merge_cells(f"B{row}:J{row}")
    _write_cell(ws, row, 2, "▶ SECTION 3. 입찰가 상향 시 순위 개선 가능 키워드",
                bold=True, bg="1F4E79", font_color="FFFFFF", font_size=11, align_h="left")
    ws.row_dimensions[row].height = 20
    row += 1
    ws.merge_cells(f"B{row}:J{row}")
    _write_cell(ws, row, 2,
                "현재 집행 중인 키워드 중 입찰가를 올리면 순위 개선 효과가 있는 키워드입니다.",
                font_size=9, font_color="595959", italic=True, align_h="left")
    row += 1

    if not bid_up_rows:
        ws.merge_cells(f"B{row}:J{row}")
        if bid_up_status == "all_optimal":
            _write_cell(ws, row, 2,
                        "✅ 현재 운영 키워드가 모두 최적 순위(1위)에 도달해 있습니다. "
                        "예산 확장 시 SECTION 2의 신규 키워드 진입을 통해 추가 트래픽을 확보하세요.",
                        font_color="1F7F3F", align_h="left", font_size=10)
        else:
            _write_cell(ws, row, 2, "입찰가 상향 추천 키워드가 없습니다.", font_color="888888", align_h="left")
        ws.row_dimensions[row].height = 16
        row += 2
    else:
        # 헤더
        for col, label in [(2,"키워드"), (3,"구분"),
                            (4,"현재 PC입찰가"), (5,"현재 MO입찰가"), (6,"현재 PC순위"), (7,"현재 MO순위"),
                            (8,"권장 PC입찰가"), (9,"권장 MO입찰가"), (10,"예상 PC순위"), (11,"예상 MO순위")]:
            _write_cell(ws, row, col, label, bold=True, bg=COLOR_HEADER_BG)
        row += 1

        for kw_row in sorted(bid_up_rows, key=lambda x: x.get("category","")):
            cat = kw_row.get("category","")
            improved = kw_row.get("rank_pc", 5) < (kw_row.get("cur_pc_rank") or 5)
            bg = "E8F5E9" if improved else None
            _write_cell(ws, row, 2, kw_row.get("keyword",""),         bg=bg, align_h="left")
            _write_cell(ws, row, 3, cat,                               bg=bg)
            _write_cell(ws, row, 4, kw_row.get("cur_pc_bid", 0),      bg=bg, num_fmt="#,##0")
            _write_cell(ws, row, 5, kw_row.get("cur_mo_bid", 0),      bg=bg, num_fmt="#,##0")
            _write_cell(ws, row, 6, kw_row.get("cur_pc_rank", "-"),   bg=bg)
            _write_cell(ws, row, 7, kw_row.get("cur_mo_rank", "-"),   bg=bg)
            _write_cell(ws, row, 8, kw_row.get("pc_bid", 0),          bg=bg, num_fmt="#,##0", bold=True)
            _write_cell(ws, row, 9, kw_row.get("mo_bid", 0),          bg=bg, num_fmt="#,##0", bold=True)
            _write_cell(ws, row, 10, kw_row.get("rank_pc", "-"),      bg=bg, bold=improved)
            _write_cell(ws, row, 11, kw_row.get("rank_mo", "-"),      bg=bg, bold=improved)
            ws.row_dimensions[row].height = 15
            row += 1


# ── 최적 효율 제안 시트 ─────────────────────────────────────────
_TIER_LABEL = {1: "★ 핵심", 2: "◆ 중요", 3: "▷ 보조"}
_TIER_COLOR = {1: "FFF2CC", 2: "E8F5E9", 3: "F3F3F3"}


def _write_optimal_sheet(ws, optimal_data, advertiser):
    """예산 제한 없는 최적 효율 제안 시트."""
    for _c, _w in {1: 3, 2: 28, 3: 14, 4: 10,
                   5: 10, 6: 10, 7: 10, 8: 10,
                   9: 10, 10: 10, 11: 10, 12: 14}.items():
        ws.column_dimensions[get_column_letter(_c)].width = _w
    ws.freeze_panes = "B5"

    optimal_rows   = optimal_data.get("optimal_rows", [])
    total_cost     = optimal_data.get("total_cost", 0)
    total_clicks   = optimal_data.get("total_clicks", 0)
    total_impr     = optimal_data.get("total_impressions", 0)

    row = 1
    # 타이틀
    ws.merge_cells(f"B{row}:L{row}")
    _write_cell(ws, row, 2, f"{advertiser}  최적 효율 제안 (전체 키워드 최적 집행 기준)",
                bold=True, font_size=14, bg=COLOR_TITLE_BG, font_color=COLOR_TITLE_FG, align_h="left")
    row += 1

    # 설명 박스
    ws.merge_cells(f"B{row}:L{row}")
    _write_cell(ws, row, 2,
                "선별된 전체 키워드를 예산 제한 없이 최적 순위(입찰가 업그레이드 포함)로 집행할 경우의 예상 성과입니다. "
                "현재 제안 예산과 무관하게 이 캠페인에서 얻을 수 있는 최대 효율 기준 총비용입니다. "
                "SA를 처음 시작하거나 예산 확대를 검토할 때 참고하세요.",
                font_size=10, font_color="475569", align_h="left")
    row += 2

    # KPI 요약 행
    kpi_cols = [
        ("예상 총 비용", f"{total_cost:,}원"),
        ("예상 노출", f"{total_impr:,}회"),
        ("예상 클릭", f"{total_clicks:,}회"),
        ("키워드 수", f"{len(optimal_rows)}개"),
    ]
    for ci, (lbl, val) in enumerate(kpi_cols, start=2):
        _write_cell(ws, row,     ci*2,   lbl, bold=True, bg="1F4E79", font_color="FFFFFF", font_size=9)
        _write_cell(ws, row + 1, ci*2,   val, bold=True, bg="EFF6FF", font_size=11)
    row += 3

    # 헤더
    headers = [
        (2, "키워드"), (3, "카테고리"), (4, "우선순위"),
        (5, "PC 입찰가"), (6, "MO 입찰가"),
        (7, "PC 순위"), (8, "MO 순위"),
        (9, "PC 노출"), (10, "MO 노출"),
        (11, "PC 클릭"), (12, "MO 클릭"),
    ]
    for col, lbl in headers:
        _write_cell(ws, row, col, lbl, bold=True, bg=COLOR_HEADER_BG)
    row += 1

    # 카테고리 순서
    sorted_rows = sorted(
        optimal_rows,
        key=lambda x: (
            _cat_sort_key(x.get("category", "일반 키워드")),
            -(x.get("recommendation_score", 50) or 50),
        )
    )

    cur_cat = None
    for r in sorted_rows:
        cat   = r.get("category", "일반 키워드")
        tier  = r.get("tier", 2)
        cc    = _cat_color(cat)
        tc    = _TIER_COLOR.get(tier, "F3F3F3")
        is_fb = r.get("is_fallback", False)
        row_bg = "FFF9C4" if is_fb else cc

        # 카테고리 구분선
        if cat != cur_cat:
            cur_cat = cat
            ws.merge_cells(f"B{row}:L{row}")
            _write_cell(ws, row, 2, f"[ {cat} ]",
                        bold=True, bg=cc, font_color="1F4E79", align_h="left")
            row += 1

        pc_bid  = _safe_int(r.get("proposed_bid_pc", 0))
        mo_bid  = _safe_int(r.get("proposed_bid_mo", 0))
        pc_rank = r.get("proposed_rank_pc", "-")
        mo_rank = r.get("proposed_rank_mo", "-")
        pc_impr = _safe_int(r.get("pc_impressions", 0))
        mo_impr = _safe_int(r.get("mo_impressions", 0))
        pc_clk  = _safe_int(r.get("pc_clicks", 0))
        mo_clk  = _safe_int(r.get("mo_clicks", 0))

        kw_label = r.get("keyword", "")
        if is_fb:
            kw_label += "  *"

        _write_cell(ws, row, 2,  kw_label,                    bg=row_bg, align_h="left")
        _write_cell(ws, row, 3,  cat,                          bg=row_bg)
        _write_cell(ws, row, 4,  _TIER_LABEL.get(tier, ""),   bg=tc)
        _write_cell(ws, row, 5,  pc_bid,  num_fmt="#,##0",     bg=row_bg)
        _write_cell(ws, row, 6,  mo_bid,  num_fmt="#,##0",     bg=row_bg)
        _write_cell(ws, row, 7,  pc_rank,                      bg=row_bg)
        _write_cell(ws, row, 8,  mo_rank,                      bg=row_bg)
        _write_cell(ws, row, 9,  pc_impr, num_fmt="#,##0",     bg=row_bg)
        _write_cell(ws, row, 10, mo_impr, num_fmt="#,##0",     bg=row_bg)
        _write_cell(ws, row, 11, pc_clk,  num_fmt="#,##0",     bg=row_bg)
        _write_cell(ws, row, 12, mo_clk,  num_fmt="#,##0",     bg=row_bg)
        row += 1

    # 합계
    row += 1
    ws.merge_cells(f"B{row}:D{row}")
    _write_cell(ws, row, 2, "합계", bold=True, bg=COLOR_SUBTOTAL_BG)
    _write_cell(ws, row, 9,  total_impr,  num_fmt="#,##0", bold=True, bg=COLOR_SUBTOTAL_BG)
    _write_cell(ws, row, 11, total_clicks, num_fmt="#,##0", bold=True, bg=COLOR_SUBTOTAL_BG)

    row += 2
    ws.merge_cells(f"B{row}:L{row}")
    _write_cell(ws, row, 2,
                f"* 이 제안서 전체를 최적 순위로 집행하려면 월 약 {total_cost:,}원의 예산이 필요합니다. "
                "* 표시 키워드는 검색량이 적어 네이버 예측 데이터가 제공되지 않습니다(카테고리 평균 기준 입찰가 제안). "
                "입찰가 업그레이드 가능 키워드는 확장 효율이 높은 bid로 자동 반영됩니다.",
                font_size=9, font_color="888888", align_h="left")