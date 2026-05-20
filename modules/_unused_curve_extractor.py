import json
import math
import os
from statistics import median

import openpyxl


# =========================
# 기본 설정
# =========================
DEFAULT_CTR_CURVE = {
    1: 0.1200,
    2: 0.0900,
    3: 0.0700,
    4: 0.0500,
    5: 0.0350,
    6: 0.0250,
    7: 0.0180,
    8: 0.0130,
    9: 0.0100,
    10: 0.0080,
}

DEFAULT_CPC_RATIO = {
    1: 1.00,
    2: 0.85,
    3: 0.72,
    4: 0.60,
    5: 0.50,
    6: 0.42,
    7: 0.36,
    8: 0.31,
    9: 0.27,
    10: 0.24,
}

CACHE_DIR = os.path.join("data", "cache")
os.makedirs(CACHE_DIR, exist_ok=True)


# =========================
# 공통 유틸
# =========================
def _to_number(v):
    if v is None:
        return 0.0

    if isinstance(v, (int, float)):
        return float(v)

    s = str(v).strip().replace(",", "").replace("%", "").replace("\n", " ")

    if s in ["", "-", "None", "null", "nan"]:
        return 0.0

    try:
        return float(s)
    except Exception:
        return 0.0


def _normalize_device(v):
    s = str(v).strip().upper()
    if s in ["PC", "P"]:
        return "PC"
    if s in ["M", "MO", "MOBILE", "모바일"]:
        return "MO"
    return s


def _safe_rank(v):
    n = int(_to_number(v))
    if n < 1:
        return 1
    if n > 10:
        return 10
    return n


def _sanitize_ratio(v, fallback):
    if v is None or not isinstance(v, (int, float)) or math.isnan(v) or math.isinf(v):
        return fallback
    if v <= 0:
        return fallback
    return float(v)


def _sanitize_ctr(v, fallback):
    if v is None or not isinstance(v, (int, float)) or math.isnan(v) or math.isinf(v):
        return fallback
    if v <= 0:
        return fallback
    if v >= 1:
        return fallback
    return float(v)


def _weighted_center(values):
    """
    평균과 중앙값을 섞어서 극단값의 영향을 줄인다.
    """
    if not values:
        return 0.0

    avg = sum(values) / len(values)
    med = median(values)

    if len(values) <= 2:
        return avg

    return (avg * 0.65) + (med * 0.35)


def _ensure_full_ranks(base_map, fallback_map):
    out = {}
    last_val = None

    for rank in range(1, 11):
        if rank in base_map and base_map[rank] > 0:
            out[rank] = float(base_map[rank])
            last_val = out[rank]
        elif last_val is not None:
            out[rank] = float(last_val)
        else:
            out[rank] = float(fallback_map[rank])

    return out


def _monotonic_nonincreasing(value_map):
    """
    rank가 내려갈수록 값이 증가하지 않도록 보정.
    1위 >= 2위 >= ... >= 10위
    """
    out = {}
    prev = None

    for rank in range(1, 11):
        cur = float(value_map[rank])

        if prev is None:
            out[rank] = cur
            prev = cur
            continue

        if cur > prev:
            cur = prev

        out[rank] = cur
        prev = cur

    return out


def _blend_with_default(curve_map, default_map, sample_count_map):
    """
    샘플 수가 적은 rank는 기본곡선과 더 많이 섞는다.
    """
    blended = {}

    for rank in range(1, 11):
        raw = float(curve_map[rank])
        default = float(default_map[rank])
        sample_count = int(sample_count_map.get(rank, 0))

        if sample_count >= 12:
            weight_raw = 0.85
        elif sample_count >= 8:
            weight_raw = 0.75
        elif sample_count >= 5:
            weight_raw = 0.60
        elif sample_count >= 3:
            weight_raw = 0.45
        elif sample_count >= 1:
            weight_raw = 0.30
        else:
            weight_raw = 0.0

        blended[rank] = (raw * weight_raw) + (default * (1 - weight_raw))

    return blended


def _json_cache_path_from_file(file_path):
    base = os.path.splitext(os.path.basename(file_path))[0]
    return os.path.join(CACHE_DIR, f"{base}_curve_cache.json")


def _write_json(path, data):
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def _read_json(path):
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


# =========================
# 엑셀 파서
# =========================
def _parse_navers_xlsx(file_path):
    """
    네이버스 파일 구조 전용 파서
    구조:
    1행: 키워드 / 희망순위 / PC / ... / 모바일 / ...
    2행: 예상입찰가 / 월평균 예상 노출수 / 월평균 예상 클릭수 / 월평균 예상 광고비 / 예상 CTR ...
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    rows = []
    for r in ws.iter_rows(min_row=3, values_only=True):
        if not r or not r[0] or not r[1]:
            continue

        keyword = str(r[0]).strip()
        rank = _safe_rank(r[1])

        # PC 블록
        pc_bid = _to_number(r[2])
        pc_impr = _to_number(r[3])
        pc_click = _to_number(r[4])
        pc_cost = _to_number(r[5])
        pc_ctr_raw = _to_number(r[6])

        pc_ctr = pc_ctr_raw / 100.0 if pc_ctr_raw > 1 else pc_ctr_raw
        if pc_ctr <= 0 and pc_impr > 0:
            pc_ctr = pc_click / pc_impr

        rows.append({
            "keyword": keyword,
            "device": "PC",
            "rank": rank,
            "impressions": pc_impr,
            "clicks": pc_click,
            "ctr": pc_ctr,
            "cpc": pc_bid,   # 네이버스는 '예상입찰가'를 CPC 후보로 사용
            "cost": pc_cost,
            "source": "navers"
        })

        # 모바일 블록
        mo_bid = _to_number(r[8])
        mo_impr = _to_number(r[9])
        mo_click = _to_number(r[10])
        mo_cost = _to_number(r[11])
        mo_ctr_raw = _to_number(r[12])

        mo_ctr = mo_ctr_raw / 100.0 if mo_ctr_raw > 1 else mo_ctr_raw
        if mo_ctr <= 0 and mo_impr > 0:
            mo_ctr = mo_click / mo_impr

        rows.append({
            "keyword": keyword,
            "device": "MO",
            "rank": rank,
            "impressions": mo_impr,
            "clicks": mo_click,
            "ctr": mo_ctr,
            "cpc": mo_bid,
            "cost": mo_cost,
            "source": "navers"
        })

    return rows


def _parse_acesquare_xlsx(file_path):
    """
    에이스퀘어 파일 구조 전용 파서
    구조:
    키워드 / 구분 / 경쟁현황 / 조회수 / 순위 / 노출수 / 클릭수 / 클릭률 / 클릭비용 / 총비용
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    rows = []
    for r in ws.iter_rows(min_row=3, values_only=True):
        if not r or not r[0] or not r[4]:
            continue

        keyword = str(r[0]).strip()
        device = _normalize_device(r[1])
        rank = _safe_rank(r[4])
        impressions = _to_number(r[5])
        clicks = _to_number(r[6])
        ctr_raw = _to_number(r[7])
        cpc = _to_number(r[8])
        cost = _to_number(r[9])

        ctr = ctr_raw / 100.0 if ctr_raw > 1 else ctr_raw
        if ctr <= 0 and impressions > 0:
            ctr = clicks / impressions

        rows.append({
            "keyword": keyword,
            "device": device,
            "rank": rank,
            "impressions": impressions,
            "clicks": clicks,
            "ctr": ctr,
            "cpc": cpc,
            "cost": cost,
            "source": "acesquare"
        })

    return rows


def extract_rows_from_excel(file_path):
    name = os.path.basename(file_path)

    if "네이버스" in name:
        return _parse_navers_xlsx(file_path)

    if "에이스퀘어" in name:
        return _parse_acesquare_xlsx(file_path)

    raise ValueError(f"지원하지 않는 파일 형식입니다: {name}")


# =========================
# 곡선 생성 로직
# =========================
def _collect_rank_samples(rows, metric_key):
    rank_map = {rank: [] for rank in range(1, 11)}
    sample_count_map = {rank: 0 for rank in range(1, 11)}

    for row in rows:
        rank = _safe_rank(row.get("rank", 0))
        value = _to_number(row.get(metric_key, 0))

        if metric_key == "ctr":
            value = _sanitize_ctr(value, 0.0)
        else:
            value = _sanitize_ratio(value, 0.0)

        if value > 0:
            rank_map[rank].append(value)
            sample_count_map[rank] += 1

    return rank_map, sample_count_map


def build_ctr_curve(rows, return_meta=False):
    rank_map, sample_count_map = _collect_rank_samples(rows, "ctr")

    raw_ctr_curve = {}
    for rank in range(1, 11):
        values = rank_map[rank]
        if values:
            raw_ctr_curve[rank] = _weighted_center(values)

    raw_ctr_curve = _ensure_full_ranks(raw_ctr_curve, DEFAULT_CTR_CURVE)
    raw_ctr_curve = _blend_with_default(raw_ctr_curve, DEFAULT_CTR_CURVE, sample_count_map)
    raw_ctr_curve = _monotonic_nonincreasing(raw_ctr_curve)

    final_curve = {}
    for rank in range(1, 11):
        final_curve[rank] = round(
            _sanitize_ctr(raw_ctr_curve[rank], DEFAULT_CTR_CURVE[rank]),
            6
        )

    if return_meta:
        return final_curve, {
            "sample_count_by_rank": sample_count_map
        }

    return final_curve


def build_cpc_ratio(rows, return_meta=False):
    """
    순위별 CPC 절대값 평균 -> 1위 기준 비율로 변환
    """
    rank_map, sample_count_map = _collect_rank_samples(rows, "cpc")

    cpc_avg = {}
    for rank in range(1, 11):
        values = rank_map[rank]
        if values:
            cpc_avg[rank] = _weighted_center(values)

    cpc_avg = _ensure_full_ranks(cpc_avg, {rank: 1000.0 for rank in range(1, 11)})

    base = cpc_avg.get(1, 0.0)
    if base <= 0:
        base = 1000.0

    ratio_map = {}
    for rank in range(1, 11):
        ratio_map[rank] = cpc_avg[rank] / base

    ratio_map[1] = 1.0
    ratio_map = _blend_with_default(ratio_map, DEFAULT_CPC_RATIO, sample_count_map)
    ratio_map[1] = 1.0
    ratio_map = _monotonic_nonincreasing(ratio_map)

    final_ratio = {}
    for rank in range(1, 11):
        fallback = DEFAULT_CPC_RATIO[rank]
        final_ratio[rank] = round(
            _sanitize_ratio(ratio_map[rank], fallback),
            6
        )

    final_ratio[1] = 1.0

    if return_meta:
        return final_ratio, {
            "sample_count_by_rank": sample_count_map
        }

    return final_ratio


# =========================
# 캐시 포함 고수준 API
# =========================
def extract_curves_from_rows(rows):
    ctr_curve = build_ctr_curve(rows)
    cpc_ratio = build_cpc_ratio(rows)
    return ctr_curve, cpc_ratio


def extract_curves_from_excel(file_path, use_cache=True, save_cache=True):
    """
    기존 코드 호환용 메인 함수.
    반환:
        ctr_curve, cpc_ratio
    """
    cache_path = _json_cache_path_from_file(file_path)

    if use_cache and os.path.exists(cache_path):
        try:
            cached = _read_json(cache_path)

            ctr_curve = {
                int(k): float(v)
                for k, v in cached.get("ctr_curve", {}).items()
            }
            cpc_ratio = {
                int(k): float(v)
                for k, v in cached.get("cpc_ratio", {}).items()
            }

            if len(ctr_curve) == 10 and len(cpc_ratio) == 10:
                return ctr_curve, cpc_ratio
        except Exception:
            pass

    try:
        rows = extract_rows_from_excel(file_path)

        ctr_curve, ctr_meta = build_ctr_curve(rows, return_meta=True)
        cpc_ratio, cpc_meta = build_cpc_ratio(rows, return_meta=True)

        if save_cache:
            cache_payload = {
                "source_file": file_path,
                "ctr_curve": ctr_curve,
                "cpc_ratio": cpc_ratio,
                "meta": {
                    "row_count": len(rows),
                    "ctr_sample_count_by_rank": ctr_meta.get("sample_count_by_rank", {}),
                    "cpc_sample_count_by_rank": cpc_meta.get("sample_count_by_rank", {}),
                }
            }
            try:
                _write_json(cache_path, cache_payload)
            except Exception:
                pass

        return ctr_curve, cpc_ratio

    except Exception as e:
        if use_cache and os.path.exists(cache_path):
            try:
                cached = _read_json(cache_path)

                ctr_curve = {
                    int(k): float(v)
                    for k, v in cached.get("ctr_curve", {}).items()
                }
                cpc_ratio = {
                    int(k): float(v)
                    for k, v in cached.get("cpc_ratio", {}).items()
                }

                if len(ctr_curve) == 10 and len(cpc_ratio) == 10:
                    print(f"⚠️ 곡선 엑셀 로드 실패, 캐시 사용: {file_path} / {e}")
                    return ctr_curve, cpc_ratio
            except Exception:
                pass

        print(f"⚠️ 곡선 생성 실패, 기본값 사용: {file_path} / {e}")
        return DEFAULT_CTR_CURVE.copy(), DEFAULT_CPC_RATIO.copy()


def extract_curves_from_multiple_excels(file_paths, use_cache=True, save_cache=True):
    """
    여러 파일의 곡선을 평균 결합.
    이후 bid_simulator에서 쓰기 좋게 만든 함수.
    """
    ctr_list = []
    cpc_list = []

    for path in file_paths:
        try:
            ctr_curve, cpc_ratio = extract_curves_from_excel(
                path,
                use_cache=use_cache,
                save_cache=save_cache
            )
            ctr_list.append(ctr_curve)
            cpc_list.append(cpc_ratio)
        except Exception:
            continue

    if not ctr_list or not cpc_list:
        return DEFAULT_CTR_CURVE.copy(), DEFAULT_CPC_RATIO.copy()

    merged_ctr = {}
    merged_cpc = {}

    for rank in range(1, 11):
        ctr_vals = [m.get(rank, DEFAULT_CTR_CURVE[rank]) for m in ctr_list]
        cpc_vals = [m.get(rank, DEFAULT_CPC_RATIO[rank]) for m in cpc_list]

        merged_ctr[rank] = round(sum(ctr_vals) / len(ctr_vals), 6)
        merged_cpc[rank] = round(sum(cpc_vals) / len(cpc_vals), 6)

    merged_ctr = _monotonic_nonincreasing(_ensure_full_ranks(merged_ctr, DEFAULT_CTR_CURVE))
    merged_cpc = _monotonic_nonincreasing(_ensure_full_ranks(merged_cpc, DEFAULT_CPC_RATIO))
    merged_cpc[1] = 1.0

    return merged_ctr, merged_cpc


def save_curve_cache_from_excel(file_path):
    """
    수동으로 캐시만 생성하고 싶을 때 사용하는 보조 함수.
    """
    ctr_curve, cpc_ratio = extract_curves_from_excel(
        file_path,
        use_cache=False,
        save_cache=True
    )
    return ctr_curve, cpc_ratio