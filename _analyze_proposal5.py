import openpyxl, sys, io, json
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

wb = openpyxl.load_workbook(r'C:\Users\Administrator\Downloads\KT_proposal_20260523 (5).xlsx', data_only=True)

# ── 제안서 시트 분석 ──────────────────────────────────────────────
ws = wb['비즈메카, BIZMEK_제안서']

# 헤더 행 찾기
keywords = []
for i, row in enumerate(ws.iter_rows(values_only=True), 1):
    cat = row[2] if len(row) > 2 else None
    dev = row[3] if len(row) > 3 else None
    if cat in ('브랜드 키워드','상품 키워드','일반 키워드','경쟁사 키워드') and dev in ('PC','MO'):
        keywords.append({
            'kw': row[1], 'cat': cat, 'dev': dev,
            'bid': row[4], 'impr': row[5], 'click': row[6], 'cost': row[7],
            'rank': row[8], 'note': row[9]
        })

pc_kws = [k for k in keywords if k['dev'] == 'PC']
print(f"=== 제안서 키워드 현황 ===")
print(f"총 PC 키워드: {len(pc_kws)}개")

from collections import Counter
cat_cnt = Counter(k['cat'] for k in pc_kws)
for cat, cnt in sorted(cat_cnt.items()):
    print(f"  {cat}: {cnt}개")

# 비용 합계
total_cost = sum((k['cost'] or 0) for k in pc_kws)
total_cost_mo = sum((k['cost'] or 0) for k in keywords if k['dev'] == 'MO')
print(f"\nPC 비용합계: {total_cost:,}원")
print(f"MO 비용합계: {total_cost_mo:,}원")
print(f"PC+MO 비용합계: {total_cost + total_cost_mo:,}원")

# 노출O 클릭0 문제 키워드
print("\n=== 노출 있는데 클릭/비용 0 키워드 (PC) ===")
prob = [k for k in pc_kws if k['impr'] and k['impr'] > 0 and (not k['click'] or k['click'] == 0)]
print(f"해당 키워드: {len(prob)}개")
for k in prob:
    print(f"  [{k['cat']}] {k['kw']:30s} 노출={k['impr']:6}  클릭={k['click']}  비용={k['cost']}  순위={k['rank']}")

# fallback vs active
active = [k for k in pc_kws if k['impr'] and k['impr'] > 0]
fallback = [k for k in pc_kws if not k['impr'] or k['impr'] == 0]
has_cost = [k for k in pc_kws if k['cost'] and k['cost'] > 0]
print(f"\n=== Estimate 성공 vs Fallback ===")
print(f"노출 있음(estimate 성공): {len(active)}개")
print(f"노출 없음(fallback): {len(fallback)}개")
print(f"비용 있음: {len(has_cost)}개")

# 전체 키워드 목록 (무관련 체크)
print("\n=== 전체 키워드 목록 (PC, 비용 순) ===")
sorted_kws = sorted(pc_kws, key=lambda x: -(x['cost'] or 0))
for k in sorted_kws:
    print(f"  [{k['cat'][:4]}] {k['kw']:30s} bid={str(k['bid']):8s} impr={str(k['impr'] or ''):8s} click={str(k['click'] or ''):5s} cost={str(k['cost'] or ''):10s} rank={k['rank']}")

# ── 확장제안 시트 ──────────────────────────────────────────────
print("\n=== 확장제안 전체 ===")
ws2 = wb['비즈메카, BIZMEK_확장제안']
for i, row in enumerate(ws2.iter_rows(values_only=True), 1):
    if any(v is not None for v in row):
        print(f"행{i}: {row}")

# ── 최적효율 시트 ──────────────────────────────────────────────
print("\n=== 최적효율제안 요약 ===")
ws3 = wb['비즈메카, BIZMEK_최적효율제안']
for i, row in enumerate(ws3.iter_rows(values_only=True), 1):
    if any(v is not None for v in row):
        print(f"행{i}: {row}")
    if i > 10:
        break

# 최적효율 키워드 집계
opt_kws = []
for i, row in enumerate(ws3.iter_rows(values_only=True), 1):
    if row[1] and row[2] and row[2] in ('브랜드 키워드','상품 키워드','일반 키워드','경쟁사 키워드'):
        opt_kws.append({
            'kw': row[1], 'cat': row[2], 'priority': row[3],
            'bid_pc': row[4], 'bid_mo': row[5],
            'rank_pc': row[6], 'rank_mo': row[7],
            'impr_pc': row[8], 'impr_mo': row[9],
            'click_pc': row[10], 'click_mo': row[11],
        })
opt_total_cost_pc = sum((r.get('click_pc') or 0) * (r.get('bid_pc') or 0) for r in opt_kws)
opt_total_cost_mo = sum((r.get('click_mo') or 0) * (r.get('bid_mo') or 0) for r in opt_kws)
print(f"\n최적효율 키워드: {len(opt_kws)}개")

# 각주에서 총 비용 찾기
for i, row in enumerate(ws3.iter_rows(values_only=True), 1):
    if row[2] and '원의 예산' in str(row[2]):
        print(f"최적효율 각주: {row[2]}")

# 제안서 마지막 각주
for i, row in enumerate(ws.iter_rows(values_only=True), 1):
    if row[1] and '추정' in str(row[1]):
        print(f"\n제안서 각주: {row[1]}")

# SA 전략 메모 찾기
print("\n=== SA 전략 메모 검색 ===")
for i, row in enumerate(ws.iter_rows(values_only=True), 1):
    if row[1] and ('전략' in str(row[1]) or 'SA' in str(row[1]) or '인사이트' in str(row[1]) or '검색 패턴' in str(row[1])):
        print(f"행{i}: {row[1][:120]}")
