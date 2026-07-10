import openpyxl, sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

path = r'c:\Users\Administrator\Downloads\(주)금화_proposal_20260709.xlsx'
wb = openpyxl.load_workbook(path, data_only=True)
print('시트:', wb.sheetnames)

for sn in wb.sheetnames:
    ws = wb[sn]
    print(f'\n=== {sn} (dims={ws.dimensions}) 상단 20행 ===')
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if i > 20:
            break
        if any(v is not None for v in row):
            print(f'행{i}: {[str(v)[:60] if v else None for v in row]}')
