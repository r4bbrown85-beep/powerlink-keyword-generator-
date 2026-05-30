import openpyxl, sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

wb = openpyxl.load_workbook(r'C:\Users\Administrator\Downloads\KT_proposal_20260524.xlsx', data_only=True)
print('시트:', wb.sheetnames)

ws_prop = wb['비즈메카, BIZMEK_제안서']

# ── 제안서 상단 전체 텍스트 (SA메모 확인) ──────────────────────────────
print('\n=== 제안서 상단 텍스트 블록 (행1~50) ===')
for i, row in enumerate(ws_prop.iter_rows(values_only=True), 1):
    if i > 50:
        break
    if any(v is not None for v in row):
        print(f'행{i}: {[str(v)[:80] if v else None for v in row[:5]]}')

# ── Overview ──────────────────────────────────────────────────────────
print('\n=== Overview ===')
ws_ov = wb['Overview']
for i, row in enumerate(ws_ov.iter_rows(values_only=True), 1):
    if any(v is not None for v in row):
        print(f'행{i}: {row}')

# ── 전체 키워드 수집 ──────────────────────────────────────────────────
keywords = []
for i, row in enumerate(ws_prop.iter_rows(values_only=True), 1):
    cat = row[2] if len(row) > 2 else None
    dev = row[3] if len(row) > 3 else None
    if cat in ('브랜드 키워드','상품 키워드','일반 키워드','경쟁사 키워드') and dev in ('PC','MO'):
        keywords.append({
            'row': i, 'kw': row[1], 'cat': cat, 'dev': dev,
            'bid': row[4], 'impr': row[5], 'click': row[6],
            'cost': row[7], 'rank': row[8], 'note': row[9]
        })

pc = [k for k in keywords if k['dev'] == 'PC']
mo = [k for k in keywords if k['dev'] == 'MO']

print(f'\n=== 키워드 현황 ===')
print(f'PC {len(pc)}개 / MO {len(mo)}개 / 합계 {len(keywords)}개')

from collections import Counter
for cat, cnt in sorted(Counter(k['cat'] for k in pc).items()):
    print(f'  [{cat}] {cnt}개')

# 비용 집계
pc_cost = sum((k['cost'] or 0) for k in pc)
mo_cost = sum((k['cost'] or 0) for k in mo)
print(f'\nPC 비용: {pc_cost:,}원  MO 비용: {mo_cost:,}원  합계: {pc_cost+mo_cost:,}원')

est_ok = [k for k in pc if k['impr'] and k['impr'] > 0]
fallback = [k for k in pc if not k['impr'] or k['impr'] == 0]
print(f'Estimate 성공: {len(est_ok)}개  Fallback: {len(fallback)}개  ({len(est_ok)/len(pc)*100:.1f}% 성공)')

# ── 문제 키워드 분류 ──────────────────────────────────────────────────
print('\n=== Estimate 성공 키워드 (비용 있음) ===')
for k in sorted(est_ok, key=lambda x: -(x['cost'] or 0)):
    print(f'  [{k["cat"][:4]}] {k["kw"]:30s} bid={str(k["bid"]):8} impr={str(k["impr"] or ""):7} click={str(k["click"] or ""):5} cost={str(k["cost"] or ""):10} rank={k["rank"]}')

print('\n=== Fallback 키워드 목록 (구분별) ===')
for cat in ['브랜드 키워드','상품 키워드','일반 키워드','경쟁사 키워드']:
    fb_cat = [k for k in fallback if k['cat'] == cat]
    if fb_cat:
        print(f'  [{cat}] {len(fb_cat)}개: {", ".join(k["kw"] for k in fb_cat[:10])}{"..." if len(fb_cat)>10 else ""}')

# ── 무관련 키워드 체크 ────────────────────────────────────────────────
irrelevant_patterns = ['업무분장','민간군사','기업분析','기업비교','기업추천','기업가격',
                       '양식','서식','한비로','업무용다이어리','중소기업검색','기업검색',
                       '비즈니스솔루션추천','기업업무솔루션추천']
print('\n=== 무관련 의심 키워드 ===')
for k in pc:
    kw = str(k['kw'] or '')
    for pat in irrelevant_patterns:
        if pat in kw:
            print(f'  [{k["cat"][:4]}] {kw:30s} bid={k["bid"]}  impr={k["impr"]}  cost={k["cost"]}  [패턴: {pat}]')
            break

# ── 확장제안 ─────────────────────────────────────────────────────────
print('\n=== 확장제안 ===')
ws_exp = wb['비즈메카, BIZMEK_확장제안']
for i, row in enumerate(ws_exp.iter_rows(values_only=True), 1):
    if any(v is not None for v in row):
        print(f'행{i}: {row}')

# ── 최적효율제안 요약 ─────────────────────────────────────────────────
print('\n=== 최적효율제안 상단 ===')
ws_opt = wb['비즈메카, BIZMEK_최적효율제안']
for i, row in enumerate(ws_opt.iter_rows(values_only=True), 1):
    if any(v is not None for v in row):
        print(f'행{i}: {row}')
    if i > 8:
        break

# 최적효율 총 비용 각주 찾기
for i, row in enumerate(ws_opt.iter_rows(values_only=True), 1):
    if row[2] and '원의 예산' in str(row[2]):
        print(f'최적효율 각주: {row[2]}')
